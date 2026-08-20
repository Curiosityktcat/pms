#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把历史项目资料按项目编号精确挂到 PMS 项目资料面板。"""
import argparse
import datetime
import os
import re
import secrets
import shutil
import sqlite3
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


PROJECT_DIR_RE = re.compile(
    r"^([A-Z]{2,}[A-Z0-9]*(?:-[A-Z0-9]+)*-?\d{4,})\s+(.+)$",
    re.IGNORECASE,
)
# 与 project_monitor_api.py 保持同一口径；WPS 是历史资料中的正常文档，不能误报为坏文件。
ALLOWED_EXT = {
    ".pdf", ".doc", ".docx", ".wps", ".xls", ".xlsx", ".ppt", ".pptx",
    ".jpg", ".jpeg", ".png", ".gif", ".zip", ".rar", ".txt", ".csv",
}
UPLOAD_ROOT = Path(__file__).resolve().parents[2] / "uploads" / "project_files"


@dataclass
class ArchiveFile:
    path: Path
    original_name: str
    folder: str
    size: int


@dataclass
class ProjectFolder:
    path: Path
    number: str
    folder_name: str
    parent: str
    files: list[ArchiveFile] = field(default_factory=list)


def normalize_number(value):
    """编号比较只统一大小写并去尾横线，不改中间字符，避免把资料错绑到相似项目。"""
    return str(value or "").strip().upper().rstrip("-")


def parse_args():
    parser = argparse.ArgumentParser(description="导入 PMS 历史项目资料")
    parser.add_argument("--root", required=True, help="历史资料根目录")
    parser.add_argument("--dry-run", action="store_true", help="只统计，不复制、不写数据库")
    parser.add_argument("--limit", type=int, help="最多导入（或预计导入）多少个新文件")
    parser.add_argument("--report", help="缺件表 xlsx 输出路径")
    args = parser.parse_args()
    if args.limit is not None and args.limit < 1:
        parser.error("--limit 必须是大于 0 的整数")
    return args


def open_database():
    raw = (os.environ.get("PMS_DB_PATH") or "").strip()
    if not raw:
        raise RuntimeError("未设置 PMS_DB_PATH；为避免误连生产库，脚本不提供默认数据库路径")
    path = Path(raw).expanduser().resolve()
    if not path.is_file():
        raise RuntimeError(f"PMS_DB_PATH 指向的数据库不存在：{path}")
    # mode=rw 防止路径拼错时 sqlite 自动创建一个空库，让人工误以为导入成功。
    conn = sqlite3.connect(path.as_uri() + "?mode=rw", uri=True)
    conn.row_factory = sqlite3.Row
    return conn, path


def load_projects(conn):
    columns = {row[1] for row in conn.execute("PRAGMA table_info(projects)")}
    required = {"id", "number", "name", "method", "year", "agency_code"}
    missing = sorted(required - columns)
    if missing:
        raise RuntimeError("projects 表缺少字段：" + "、".join(missing))
    officer_sql = "officer" if "officer" in columns else "'' AS officer"
    # 已软删除的项目（含作废编号的墓碑记录）不接收资料，也不该出现在缺件表里——
    # 否则会催代理公司去补一个废号的材料。
    deleted_sql = ""
    if "is_deleted" in columns:
        deleted_sql = " WHERE COALESCE(is_deleted, 0) = 0"
    rows = conn.execute(
        "SELECT id, number, name, method, year, agency_code, " + officer_sql
        + " FROM projects" + deleted_sql
    ).fetchall()
    by_number = defaultdict(list)
    for row in rows:
        number = normalize_number(row["number"])
        if number:
            by_number[number].append(row)
    return rows, by_number


def project_file_columns(conn):
    columns = {row[1] for row in conn.execute("PRAGMA table_info(project_files)")}
    required = {"project_id", "original_name", "saved_name", "size", "uploaded_by", "uploaded_at"}
    missing = sorted(required - columns)
    if missing:
        raise RuntimeError("project_files 表缺少字段：" + "、".join(missing))
    return columns


def scan_project_files(project_path, errors, exclude_dirs=()):
    """收集项目文件夹下的文件；exclude_dirs 里的子树是别的项目，整棵剪掉。"""
    files = []
    excluded = {Path(d) for d in exclude_dirs}

    def onerror(error):
        errors.append(f"{getattr(error, 'filename', project_path)}：{error}")

    for current, dirs, names in os.walk(project_path, onerror=onerror, followlinks=False):
        dirs.sort()
        names.sort()
        dirs[:] = [d for d in dirs if (Path(current) / d) not in excluded]
        current_path = Path(current)
        relative_dir = current_path.relative_to(project_path)
        folder = "" if relative_dir == Path(".") else relative_dir.as_posix()
        for name in names:
            path = current_path / name
            try:
                size = path.stat().st_size
            except OSError as error:
                errors.append(f"{path}：读取大小失败：{error}")
                continue
            files.append(ArchiveFile(path=path, original_name=name, folder=folder, size=size))
    return files


def discover_project_folders(root, errors, known_numbers=()):
    """找出所有项目文件夹。

    项目文件夹里偶尔还套着另一个带编号的文件夹，有两种情况要分开处理：
      · 里层编号在 PMS 里是**另一个**项目 → 它得独立成项目，否则资料会挂错人；
      · 里层编号不在 PMS，或就是外层同一个编号（第2次采购之类）→ 并进外层当子目录，
        这样材料不至于丢掉。
    """
    known = {normalize_number(n) for n in known_numbers}
    candidates = []
    def onerror(error):
        errors.append(f"{getattr(error, 'filename', root)}：扫描目录失败：{error}")

    for current, dirs, _names in os.walk(root, topdown=True, onerror=onerror, followlinks=False):
        dirs.sort()
        current_path = Path(current)
        if current_path == root:
            continue
        match = PROJECT_DIR_RE.match(current_path.name)
        if match:
            candidates.append((current_path, normalize_number(match.group(1))))

    # 只有「里层编号在 PMS 里且与外层不同」才算独立项目
    by_path = dict(candidates)
    independent = []
    for path, number in candidates:
        outer = None
        parent = path.parent
        while True:
            if parent in by_path:
                outer = parent
                break
            if parent == root or parent == parent.parent:
                break
            parent = parent.parent
        if outer is None:
            independent.append((path, number))
        elif number in known and number != by_path[outer]:
            independent.append((path, number))
    independent_paths = {p for p, _ in independent}

    found = []
    for path, number in independent:
        nested = {c for c in independent_paths if c != path and path in c.parents}
        parent_path = path.parent.relative_to(root)
        parent = root.name if parent_path == Path(".") else parent_path.as_posix()
        found.append(ProjectFolder(
            path=path,
            number=number,
            folder_name=path.name,
            parent=parent,
            files=scan_project_files(path, errors, exclude_dirs=nested),
        ))
    return found


def load_existing(conn, has_folder):
    folder_sql = "COALESCE(folder, '')" if has_folder else "''"
    existing = defaultdict(set)
    for row in conn.execute(
        f"SELECT project_id, {folder_sql} AS folder, original_name, size FROM project_files"
    ):
        existing[row["project_id"]].add((
            row["folder"] or "", row["original_name"] or "", int(row["size"] or 0)
        ))
    return existing


def insert_one(conn, project_id, archive_file, destination, saved_name):
    now = datetime.datetime.now().isoformat(timespec="seconds")
    # 每个文件单独提交稍慢，但机器或脚本中断时不会留下整批只有磁盘文件、没有数据库记录的孤件。
    try:
        conn.execute(
            "INSERT INTO project_files "
            "(project_id, original_name, saved_name, folder, size, uploaded_by, uploaded_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (project_id, archive_file.original_name, saved_name, archive_file.folder,
             archive_file.size, "历史资料导入", now),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        try:
            destination.unlink()
        except OSError:
            pass
        raise


def process_files(conn, matched, existing, dry_run=False, limit=None):
    imported = 0
    skipped = 0
    shortcut_skipped = 0
    errors = []
    covered = defaultdict(lambda: {"files": 0, "folders": set()})
    covered_signatures = defaultdict(set)
    limit_reached = False

    def mark_covered(project_id, archive_file, signature):
        # 两处历史目录若碰巧放了同名同大小文件，幂等规则只会落一行；
        # 汇总表也按最终能看到的记录数算，不能把同一签名重复计数。
        if signature in covered_signatures[project_id]:
            return
        covered_signatures[project_id].add(signature)
        covered[project_id]["files"] += 1
        if archive_file.folder:
            covered[project_id]["folders"].add(archive_file.folder)

    for folder, project in matched:
        project_id = project["id"]
        for archive_file in folder.files:
            ext = archive_file.path.suffix.lower()
            if ext == ".lnk":
                shortcut_skipped += 1
                continue
            if ext not in ALLOWED_EXT:
                errors.append(
                    f"{archive_file.path}：不支持的格式 {ext or '无扩展名'}"
                )
                continue
            signature = (archive_file.folder, archive_file.original_name, archive_file.size)
            if signature in existing[project_id]:
                skipped += 1
                mark_covered(project_id, archive_file, signature)
                continue
            if limit is not None and imported >= limit:
                limit_reached = True
                continue

            if dry_run:
                imported += 1
                existing[project_id].add(signature)
                mark_covered(project_id, archive_file, signature)
                continue

            destination_dir = UPLOAD_ROOT / str(project_id)
            destination = None
            try:
                destination_dir.mkdir(parents=True, exist_ok=True)
                stamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                saved_name = f"{stamp}_{secrets.token_hex(4)}{ext}"
                destination = destination_dir / saved_name
                shutil.copy2(archive_file.path, destination)
                insert_one(conn, project_id, archive_file, destination, saved_name)
            except Exception as error:                         # noqa: BLE001
                # copy2 自身也可能在写到一半后失败；残件没有数据库记录，必须当场清掉，
                # 否则重跑虽不会展示重复行，磁盘上却会不断积累孤件。
                if destination is not None and destination.exists():
                    try:
                        destination.unlink()
                    except OSError:
                        pass
                errors.append(f"{archive_file.path}：复制或写库失败：{error}")
                continue
            existing[project_id].add(signature)
            imported += 1
            mark_covered(project_id, archive_file, signature)

    return imported, skipped, shortcut_skipped, errors, covered, limit_reached


def write_report(path, projects, matched, unmatched, covered):
    matched_numbers = {folder.number for folder, _project in matched}
    wb = Workbook()
    ws_missing = wb.active
    ws_missing.title = "无资料项目"
    ws_unmatched = wb.create_sheet("编号对不上")
    ws_summary = wb.create_sheet("已导入汇总")

    def add_header(ws, headers, widths):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, index).column_letter].width = width

    add_header(ws_missing,
               ["项目编号", "项目名称", "采购方式", "采购年度", "代理机构代码", "经办人"],
               [24, 48, 18, 12, 16, 14])
    missing_projects = [
        row for row in projects
        if normalize_number(row["number"]) and normalize_number(row["number"]) not in matched_numbers
    ]
    missing_projects.sort(key=lambda row: (
        row["agency_code"] or "", normalize_number(row["number"]), row["id"]
    ))
    for row in missing_projects:
        ws_missing.append([row["number"] or "", row["name"] or "", row["method"] or "",
                           row["year"] or "", row["agency_code"] or "", row["officer"] or ""])

    add_header(ws_unmatched,
               ["文件夹里的编号", "文件夹名", "文件数", "所在上级目录"],
               [26, 58, 12, 42])
    for folder in sorted(unmatched, key=lambda item: (item.number, item.parent, item.folder_name)):
        ws_unmatched.append([folder.number, folder.folder_name, len(folder.files), folder.parent])

    add_header(ws_summary,
               ["项目编号", "项目名称", "导入文件数", "涉及子文件夹数"],
               [26, 52, 14, 18])
    matched_projects = {}
    for _folder, project in matched:
        matched_projects[project["id"]] = project
    for project_id, project in sorted(
            matched_projects.items(), key=lambda item: normalize_number(item[1]["number"])):
        info = covered[project_id]
        ws_summary.append([project["number"] or "", project["name"] or "",
                           info["files"], len(info["folders"])])

    output = Path(path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def print_results(folders, matched, unmatched, imported, skipped, shortcut_skipped,
                  errors, dry_run, limit_reached):
    matched_project_count = len({project["id"] for _folder, project in matched})
    print(f"识别出带编号的项目文件夹：{len(folders)} 个")
    print(f"编号匹配上的项目：{matched_project_count} 个，共 {sum(len(f.files) for f, _ in matched)} 个文件")
    print(f"{'预计导入' if dry_run else '导入'}文件数：{imported}")
    print(f"幂等跳过数：{skipped}")
    print(f"快捷方式跳过数：{shortcut_skipped}")
    if limit_reached:
        print("已达到 --limit，剩余可导入文件未处理")
    print(f"编号对不上的文件夹：{len(unmatched)} 个")
    if unmatched:
        for folder in sorted(unmatched, key=lambda item: (item.number, item.parent)):
            print(f"  - {folder.number}｜{folder.folder_name}｜{len(folder.files)} 个文件｜{folder.parent}")
    else:
        print("  （无）")
    print(f"出错的文件：{len(errors)} 个")
    if errors:
        for error in errors:
            print(f"  - {error}")
    else:
        print("  （无）")


def main():
    args = parse_args()
    root = Path(args.root).expanduser().resolve()
    if not root.is_dir():
        raise SystemExit(f"资料根目录不存在或不是目录：{root}")

    conn = None
    try:
        conn, db_path = open_database()
        projects, projects_by_number = load_projects(conn)
        columns = project_file_columns(conn)
        if "folder" not in columns and not args.dry_run:
            raise RuntimeError(
                "project_files 表还没有 folder 列，请先用新代码启动一次 PMS 完成自动补列"
            )
        scan_errors = []
        folders = discover_project_folders(root, scan_errors,
                                           known_numbers=projects_by_number.keys())
        matched = []
        unmatched = []
        for folder in folders:
            candidates = projects_by_number.get(folder.number, [])
            if len(candidates) == 1:
                matched.append((folder, candidates[0]))
            elif not candidates:
                unmatched.append(folder)
            else:
                # 规范化后仍重复就不导；无法唯一确定时宁可漏绑，也不能猜一个项目写进去。
                scan_errors.append(
                    f"{folder.path}：PMS 中编号 {folder.number} 规范化后对应多个项目，已跳过"
                )

        existing = load_existing(conn, "folder" in columns)
        imported, skipped, shortcut_skipped, run_errors, covered, limit_reached = process_files(
            conn, matched, existing, dry_run=args.dry_run, limit=args.limit
        )
        errors = scan_errors + run_errors
        if args.report:
            report_path = write_report(args.report, projects, matched, unmatched, covered)
            print(f"缺件表已生成：{report_path}")
        print(f"数据库：{db_path}")
        print(f"资料根目录：{root}")
        print_results(folders, matched, unmatched, imported, skipped, shortcut_skipped,
                      errors, args.dry_run, limit_reached)
    except (RuntimeError, sqlite3.Error) as error:
        raise SystemExit(f"导入失败：{error}") from error
    finally:
        if conn is not None:
            conn.close()


if __name__ == "__main__":
    main()

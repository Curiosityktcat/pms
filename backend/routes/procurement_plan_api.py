# -*- coding: utf-8 -*-
"""采购计划池接口：归口科室的年度采购计划，按年度/科室/分类筛选，可与采购项目挂钩。"""
import datetime
import io
import json
import os

from flask import Blueprint, request, session, jsonify, send_file

from models import db
from models.project import Project
from models.procurement_plan import (ProcurementPlan, ProcurementPlanAttachment,
                                     NOT_PROCURED)
from routes.utils import login_required

bp = Blueprint("procurement_plan", __name__, url_prefix="/api/procurement-plans")

UPLOAD_ROOT = "/home/huangxb/pms/uploads/procurement_plan"
ALLOWED = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".jpg", ".jpeg", ".png",
           ".gif", ".zip", ".rar", ".txt"}


def _now():
    return datetime.datetime.now().isoformat(timespec="seconds")


def _my_dept_names():
    """科室账号的本科室名字集合（含曾用名）；非科室账号返回 None 表示不限。"""
    from services.dept_scope import is_dept_role, current_dept_code
    if not is_dept_role():
        return None
    from services.dept import dept_names
    return set(dept_names(current_dept_code()) or [])


def _is_mine(plan) -> bool:
    """这条计划是不是本科室的。归口或需求科室命中其一即可。"""
    names = _my_dept_names()
    if names is None:
        return True
    return (plan.dept or "") in names or (plan.demand_dept or "") in names


def _can_edit(plan=None):
    """谁能改计划。

    原来只有采购部能改，科室连自己报的计划都动不了——那「小团队搬进 PMS」
    就没搬成，还是采购部在两边录（2026-08-18 用户：「和小团队的数据结合，没有做到位」）。
    现在归口科室可以维护**本科室**的计划，越界的照样拒。
    """
    role = session.get("role")
    if role in ("officer", "assistant", "leader", "admin"):
        return True
    from services.dept_scope import is_dept_role
    if is_dept_role():
        from services.permission import get_user_perms
        if "procurement-plan" not in set(get_user_perms(session.get("user", ""), role)):
            return False
        return True if plan is None else _is_mine(plan)
    return False


def _guard_read(plan_id):
    """读单条计划及其附件时的科室收口——列表按科室筛过了，但改 URL 直接点 id 不能绕过。"""
    row = db.session.get(ProcurementPlan, plan_id)
    if not row:
        return None, (jsonify({"ok": False, "error": "计划不存在"}), 404)
    if not _is_mine(row):
        return None, (jsonify({"ok": False, "error": "无权查看该计划"}), 403)
    return row, None


def _load_and_guard(plan_id, what):
    """先取记录再判权限：科室只能动本科室的计划。

    原来五个写接口都是先 _can_edit() 再取记录，而那时还不知道这条计划属于谁——
    科室只要有计划池权限就能改全院任何一条。
    """
    row = db.session.get(ProcurementPlan, plan_id)
    if not row:
        return None, (jsonify({"ok": False, "error": "计划不存在"}), 404)
    if not _can_edit(row):
        return None, (jsonify({"ok": False, "error": f"无权{what}（只能操作本科室的计划）"}), 403)
    return row, None


@bp.route("/meta", methods=["GET"])
@login_required
def meta():
    """筛选项都从真实数据里取，别写死——各科室的分类口径本来就不统一。"""
    rows = db.session.execute(db.select(ProcurementPlan)).scalars().all()

    def uniq(attr):
        return sorted({(getattr(r, attr) or "").strip() for r in rows} - {""})

    return jsonify({"ok": True, "data": {
        "years": sorted({r.year for r in rows if r.year}, reverse=True),
        "depts": uniq("dept"),
        "categories": uniq("category"),
        "categories2": uniq("category2"),
        "methods": uniq("method"),
        "org_forms": uniq("org_form"),
        "statuses": uniq("status"),
        "demand_types": uniq("demand_type"),
        "not_procured": list(NOT_PROCURED),
        "can_edit": _can_edit(),
        "total": len(rows),
    }})


@bp.route("", methods=["GET"])
@login_required
def list_plans():
    a = request.args
    conds = []
    if a.get("year"):
        conds.append(ProcurementPlan.year == int(a["year"]))
    for field in ("dept", "category", "category2", "method", "org_form",
                  "status", "demand_type"):
        if a.get(field):
            conds.append(getattr(ProcurementPlan, field) == a[field])

    # 科室账号只看本科室的计划：这个模块放开给科室之后，不收口就等于把
    # 全院各科室的年度计划和预算摊给每个科室看。
    names = _my_dept_names()
    if names is not None:
        if not names:
            return jsonify({"ok": True, "data": [], "total": 0})
        conds.append(db.or_(ProcurementPlan.dept.in_(names),
                            ProcurementPlan.demand_dept.in_(names)))

    rows = db.session.execute(
        db.select(ProcurementPlan).where(*conds)
    ).scalars().all()

    kw = (a.get("keyword") or "").strip()
    if kw:
        rows = [r for r in rows if kw in (r.name or "") or kw in (r.plan_number or "")
                or kw in (r.package_no or "") or kw in (r.note or "")]

    # linked=1 只看已关联项目的，0 只看没关联的
    if a.get("linked") == "1":
        rows = [r for r in rows if r.project_id]
    elif a.get("linked") == "0":
        rows = [r for r in rows if not r.project_id]

    # 默认把「已合并/已集采/延期合并」这类不会走到采购部的藏起来，
    # 它们只是科室台账里的痕迹，混在待立项里会让人误以为还有一堆活没干
    if a.get("include_closed") != "1":
        rows = [r for r in rows if (r.status or "") not in NOT_PROCURED]

    projs = {}
    pids = [r.project_id for r in rows if r.project_id]
    if pids:
        projs = {p.id: p for p in db.session.execute(
            db.select(Project).where(Project.id.in_(pids))).scalars().all()}

    att_count = {}
    for pid, n in db.session.execute(
        db.select(ProcurementPlanAttachment.plan_id,
                  db.func.count(ProcurementPlanAttachment.id))
        .group_by(ProcurementPlanAttachment.plan_id)
    ).all():
        att_count[pid] = n

    out = []
    for r in rows:
        d = r.to_dict(projs.get(r.project_id))
        d["attachment_count"] = att_count.get(r.id, 0)
        out.append(d)
    # 有编号的按编号排，其余按科室+名称，保证顺序稳定
    out.sort(key=lambda d: (not d["plan_number"], d["plan_number"], d["dept"], d["name"]))
    return jsonify({"ok": True, "data": out, "total": len(out)})


@bp.route("/stats", methods=["GET"])
@login_required
def stats():
    """给页面顶部的概览卡片：多少条、多少已立项、多少不会进采购部、预算合计。"""
    year = request.args.get("year")
    conds = [ProcurementPlan.year == int(year)] if year else []
    rows = db.session.execute(db.select(ProcurementPlan).where(*conds)).scalars().all()
    closed = [r for r in rows if (r.status or "") in NOT_PROCURED]
    live = [r for r in rows if r not in closed]
    linked = [r for r in live if r.project_id]
    by_dept = {}
    for r in live:
        k = r.dept or "（未填）"
        by_dept[k] = by_dept.get(k, 0) + 1
    return jsonify({"ok": True, "data": {
        "total": len(rows),
        "live": len(live),
        "closed": len(closed),
        "linked": len(linked),
        "unlinked": len(live) - len(linked),
        "budget_sum": round(sum(r.budget or 0 for r in live), 2),
        "by_dept": sorted(by_dept.items(), key=lambda x: -x[1]),
    }})


@bp.route("/<int:plan_id>", methods=["PUT"])
@login_required
def update_plan(plan_id):
    row, err = _load_and_guard(plan_id, "修改采购计划")
    if err:
        return err
    data = request.get_json(force=True) or {}
    for k in ("name", "package_no", "plan_number", "dept", "demand_dept", "org_form",
              "method", "deadline", "qty", "unit", "category", "category2",
              "demand_type", "status", "note"):
        if k in data:
            setattr(row, k, (data[k] or "").strip() if isinstance(data[k], str) else data[k])
    for k in ("budget", "price_limit"):
        if k in data:
            try:
                setattr(row, k, float(data[k] or 0))
            except (TypeError, ValueError):
                pass
    row.updated_at = _now()
    db.session.commit()
    return jsonify({"ok": True, "data": row.to_dict()})


@bp.route("/<int:plan_id>/link", methods=["POST"])
@login_required
def link_project(plan_id):
    """把计划挂到正式采购项目上。

    只接受人工点选或按编号定位——名称自动匹配错绑的代价比人工点一下大得多
    （科室写「雾化器【6元版本】」，立项写「2026年气动雾化吸入器采购项目」）。
    """
    row, err = _load_and_guard(plan_id, "关联项目")
    if err:
        return err
    data = request.get_json(force=True) or {}
    pid = data.get("project_id")
    number = (data.get("project_number") or "").strip()

    proj = None
    if pid:
        proj = db.session.get(Project, int(pid))
    elif number:
        proj = db.session.execute(
            db.select(Project).filter_by(number=number)).scalars().first()
    if not proj:
        return jsonify({"ok": False, "error": "找不到该采购项目"}), 404

    other = db.session.execute(db.select(ProcurementPlan).where(
        ProcurementPlan.project_id == proj.id,
        ProcurementPlan.id != plan_id)).scalars().first()
    if other:
        return jsonify({"ok": False,
                        "error": f"该项目已关联计划「{other.name}」，请先解除"}), 400

    row.project_id = proj.id
    row.plan_number = proj.number or row.plan_number
    row.linked_by = session.get("display_name", "")
    row.linked_at = _now()
    row.updated_at = row.linked_at
    db.session.commit()
    return jsonify({"ok": True, "data": row.to_dict(proj), "message": "已关联采购项目"})


@bp.route("/<int:plan_id>/link", methods=["DELETE"])
@login_required
def unlink_project(plan_id):
    row, err = _load_and_guard(plan_id, "解除关联")
    if err:
        return err
    row.project_id = None
    row.linked_by = ""
    row.linked_at = ""
    row.updated_at = _now()
    db.session.commit()
    return jsonify({"ok": True, "message": "已解除关联"})


@bp.route("/<int:plan_id>/candidates", methods=["GET"])
@login_required
def candidates(plan_id):
    """关联时的候选项目：按关键词给提示，只提示不自动绑。"""
    row, err = _guard_read(plan_id)
    if err:
        return err
    kw = (request.args.get("keyword") or "").strip()
    taken = {p.project_id for p in db.session.execute(
        db.select(ProcurementPlan).where(ProcurementPlan.project_id.isnot(None))
    ).scalars().all()} - {row.project_id}

    _stmt = db.select(Project).where(
        db.or_(Project.is_deleted == 0, Project.is_deleted.is_(None)),
        db.or_(Project.is_draft == 0, Project.is_draft.is_(None)),
    )
    # 科室账号的候选只能是本科室的项目——否则关联弹窗会把全院项目名摊出来
    from services.dept_scope import is_dept_role, scope_projects
    if is_dept_role():
        _stmt = scope_projects(_stmt)
    projs = db.session.execute(_stmt).scalars().all()

    def score(p):
        if kw:
            return 1 if (kw in (p.name or "") or kw in (p.number or "")) else 0
        # 没给关键词时，用计划名里的字与项目名的重合度粗排（只做提示）
        name = (row.name or "").strip()
        if not name:
            return 0
        hit = sum(1 for ch in set(name) if ch in (p.name or ""))
        return hit / max(len(set(name)), 1)

    scored = [(score(p), p) for p in projs if p.id not in taken]
    scored = [(s, p) for s, p in scored if s > 0]
    scored.sort(key=lambda x: -x[0])
    return jsonify({"ok": True, "data": [{
        "id": p.id, "number": p.number or "", "name": p.name,
        "status": p.status or "", "officer": p.officer or "",
        "match": round(float(s), 3),
    } for s, p in scored[:30]]})


# ══════════════════════════════════════════════════════════════════
# Excel 对照表：导出 → 人工填项目编号 → 传回来批量关联
#
# 用户 2026-08-18：「你可以做一个项目清单的 excel 表，然后我把小团队项目名称和
# 项目编号手动关联一下，我觉得这步这样做效率更高」。
# 在界面上一条条点太慢（365 条），Excel 里对着两列看、复制粘贴快得多。
# 仍然是人工关联——不做任何自动匹配，只是把「点」换成了「填」。
# ══════════════════════════════════════════════════════════════════

# 「请填这里」那一列的表头，导入时靠它定位，改了表头就对不上了
COL_FILL = "项目编号（请填这里）"


def _plan_rows_for_export():
    """要导出的计划：本科室（科室账号）或全部（采购部），已结案的不导。"""
    conds = [ProcurementPlan.status.notin_(NOT_PROCURED)]
    names = _my_dept_names()
    if names is not None:
        if not names:
            return []
        conds.append(db.or_(ProcurementPlan.dept.in_(names),
                            ProcurementPlan.demand_dept.in_(names)))
    rows = db.session.execute(db.select(ProcurementPlan).where(*conds)).scalars().all()
    # 没关联的排前面——那才是要干的活
    rows.sort(key=lambda r: (r.project_id is not None, r.dept or "", r.name or ""))
    return rows


@bp.route("/link-sheet", methods=["GET"])
@login_required
def link_sheet():
    """导出对照表。两个工作表：第一张填编号，第二张是项目清单供查号。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows = _plan_rows_for_export()
    pids = [r.project_id for r in rows if r.project_id]
    projs = {p.id: p for p in db.session.execute(
        db.select(Project).where(Project.id.in_(pids))).scalars()} if pids else {}

    wb = Workbook()
    ws = wb.active
    ws.title = "计划对照表"
    head = ["计划ID", "计划名称", "归口科室", "需求科室", "预算（元）", "采购期限",
            "状态", "当前已关联项目", COL_FILL]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A73E8")
        c.alignment = Alignment(vertical="center", wrap_text=True)
    fill_col = len(head)
    ws.cell(1, fill_col).fill = PatternFill("solid", fgColor="F9AB00")

    for r in rows:
        pr = projs.get(r.project_id)
        ws.append([r.id, r.name or "", r.dept or "", r.demand_dept or "",
                   r.budget or 0, r.deadline or "", r.status or "",
                   f"{pr.number or ''}　{pr.name or ''}" if pr else "",
                   pr.number if pr else ""])
    for i, w in enumerate((8, 42, 14, 14, 13, 14, 10, 40, 22), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # 第二张：项目清单，供查编号
    ws2 = wb.create_sheet("项目清单（查编号用）")
    stmt = db.select(Project).where(
        db.or_(Project.is_deleted == 0, Project.is_deleted.is_(None)),
        db.or_(Project.is_draft == 0, Project.is_draft.is_(None)),
    )
    from services.dept_scope import is_dept_role, scope_projects
    if is_dept_role():
        stmt = scope_projects(stmt)
    ws2.append(["项目编号", "项目名称", "归口科室", "需求科室", "经办人",
                "采购方式", "预算（元）", "状态", "已被哪条计划占用"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="34A853")
    taken = {}
    for pl in db.session.execute(db.select(ProcurementPlan).where(
            ProcurementPlan.project_id.isnot(None))).scalars():
        taken[pl.project_id] = pl.name or f"计划#{pl.id}"
    for pr in db.session.execute(stmt).scalars():
        ws2.append([pr.number or "", pr.name or "", pr.manage_dept or "",
                    pr.demand_dept or "", pr.officer or "", pr.method or "",
                    pr.amount, pr.status or "", taken.get(pr.id, "")])
    for i, w in enumerate((22, 46, 14, 14, 10, 16, 13, 10, 30), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # 第三张：怎么填
    ws3 = wb.create_sheet("怎么填")
    for line in [
        ["怎么用这张表"],
        [""],
        ["1", "打开「计划对照表」，未关联的计划排在最前面。"],
        ["2", f"在最后一列「{COL_FILL}」里填上对应的采购项目编号。"],
        ["3", "编号去「项目清单（查编号用）」那张表里找，复制粘贴过来即可。"],
        ["4", "填完保存，回 PMS 的「1.0 采购计划池」页面点「导入对照表」传上去。"],
        [""],
        ["注意"],
        ["", "· 不用每行都填，填了的才会处理，空着的一律不动。"],
        ["", "· 「计划ID」那一列不要改，系统靠它认是哪条计划。"],
        ["", "· 一个采购项目只能被一条计划关联；重复填会在导入结果里告诉你。"],
        ["", "· 想解除某条已有的关联，把最后一列清空是不管用的（空=不处理），"],
        ["", "  请在页面上点「解除关联」。"],
        ["", "· 导入会先给你一份预览，确认无误再真正写入。"],
    ]:
        ws3.append(line)
    ws3["A1"].font = Font(bold=True, size=14)
    ws3["A8"].font = Font(bold=True)
    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 76

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"采购计划对照表_{datetime.date.today().isoformat()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/link-import", methods=["POST"])
@login_required
def link_import():
    """导入填好的对照表。dry_run=1 只预览不写库。"""
    if not _can_edit():
        return jsonify({"ok": False, "error": "无权关联项目"}), 403
    f = (request.files.get("file") or request.files.get("files"))
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "没有收到文件"}), 400
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"ok": False, "error": "请上传 .xlsx 文件（就是导出的那张表填好）"}), 400
    dry = str(request.form.get("dry_run", "")).lower() in ("1", "true", "yes")

    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:                                   # noqa: BLE001
        return jsonify({"ok": False, "error": f"这个文件打不开：{e}"}), 400
    ws = wb["计划对照表"] if "计划对照表" in wb.sheetnames else wb.worksheets[0]

    head = [str(c.value or "").strip() for c in ws[1]]
    if "计划ID" not in head or COL_FILL not in head:
        return jsonify({"ok": False,
                        "error": f"表头对不上，缺「计划ID」或「{COL_FILL}」。"
                                 "请用系统导出的那张表来填，别自己新建。"}), 400
    i_id, i_num = head.index("计划ID"), head.index(COL_FILL)

    # 一次性把要用的项目查出来，别在循环里逐条查库
    allowed = _my_dept_names()
    ok_rows, skip, errors = [], 0, []
    seen_numbers = {}
    for rno in range(2, ws.max_row + 1):
        raw_id = ws.cell(rno, i_id + 1).value
        raw_num = ws.cell(rno, i_num + 1).value
        number = str(raw_num or "").strip()
        if not number:
            skip += 1
            continue
        try:
            plan_id = int(raw_id)
        except (TypeError, ValueError):
            errors.append(f"第 {rno} 行：计划ID「{raw_id}」不是数字，跳过")
            continue
        plan = db.session.get(ProcurementPlan, plan_id)
        if plan is None:
            errors.append(f"第 {rno} 行：找不到计划 #{plan_id}")
            continue
        if allowed is not None and not _is_mine(plan):
            errors.append(f"第 {rno} 行：计划「{plan.name}」不属于本科室")
            continue
        proj = db.session.execute(
            db.select(Project).filter_by(number=number)).scalars().first()
        if proj is None:
            errors.append(f"第 {rno} 行：找不到项目编号「{number}」")
            continue
        if number in seen_numbers:
            errors.append(f"第 {rno} 行：编号「{number}」在表里出现了不止一次"
                          f"（另一处是第 {seen_numbers[number]} 行）")
            continue
        seen_numbers[number] = rno
        if plan.project_id == proj.id:
            skip += 1                       # 已经是这个关联，不算改动
            continue
        other = db.session.execute(db.select(ProcurementPlan).where(
            ProcurementPlan.project_id == proj.id,
            ProcurementPlan.id != plan_id)).scalars().first()
        if other:
            errors.append(f"第 {rno} 行：项目「{number}」已被计划「{other.name}」占用")
            continue
        ok_rows.append({"row": rno, "plan_id": plan_id, "plan_name": plan.name or "",
                        "project_id": proj.id, "project_number": number,
                        "project_name": proj.name or "",
                        "was_linked": plan.project_id is not None})

    if not dry:
        who = session.get("display_name", "")
        for item in ok_rows:
            plan = db.session.get(ProcurementPlan, item["plan_id"])
            plan.project_id = item["project_id"]
            plan.plan_number = item["project_number"] or plan.plan_number
            plan.linked_by = who
            plan.linked_at = _now()
            plan.updated_at = plan.linked_at
        db.session.commit()

    return jsonify({"ok": True, "dry_run": dry, "data": ok_rows,
                    "will_link": len(ok_rows), "skipped": skip, "errors": errors,
                    "message": (f"预览：可关联 {len(ok_rows)} 条"
                                if dry else f"已关联 {len(ok_rows)} 条")
                               + (f"，{len(errors)} 条有问题" if errors else "")})


# ── 附件：科室需求表、办公会决议、报价单等 ─────────────────────────

@bp.route("/<int:plan_id>/attachments", methods=["GET"])
@login_required
def list_attachments(plan_id):
    _p, err = _guard_read(plan_id)
    if err:
        return err
    rows = db.session.execute(
        db.select(ProcurementPlanAttachment).filter_by(plan_id=plan_id)
        .order_by(ProcurementPlanAttachment.id)).scalars().all()
    out = []
    for a in rows:
        d = a.to_dict()
        d["exists"] = bool(a.path and os.path.exists(a.path))
        out.append(d)
    return jsonify({"ok": True, "data": out})


@bp.route("/<int:plan_id>/attachments", methods=["POST"])
@login_required
def upload_attachment(plan_id):
    plan, err = _load_and_guard(plan_id, "上传附件")
    if err:
        return err
    files = request.files.getlist("file") or request.files.getlist("files")
    if not files:
        return jsonify({"ok": False, "error": "没有收到文件"}), 400

    sub = os.path.join(UPLOAD_ROOT, str(plan_id))
    os.makedirs(sub, exist_ok=True)
    saved = []
    for f in files:
        name = os.path.basename(f.filename or "")
        if not name:
            continue
        ext = os.path.splitext(name)[1].lower()
        if ext not in ALLOWED:
            return jsonify({"ok": False, "error": f"不支持的文件类型：{ext}"}), 400
        row = ProcurementPlanAttachment(
            plan_id=plan_id, filename=name, uploaded_by=session.get("display_name", ""),
            uploaded_at=_now(), source="upload")
        db.session.add(row)
        db.session.flush()                      # 先拿到 id，用它做文件名前缀防重名覆盖
        path = os.path.join(sub, f"{row.id}_{name}")
        f.save(path)
        row.path = path
        row.size = os.path.getsize(path)
        saved.append(row)
    db.session.commit()
    return jsonify({"ok": True, "data": [r.to_dict() for r in saved],
                    "message": f"已上传 {len(saved)} 个文件"})


def _get_att(plan_id, aid):
    return db.session.execute(db.select(ProcurementPlanAttachment)
                              .filter_by(id=aid, plan_id=plan_id)).scalars().first()


@bp.route("/<int:plan_id>/attachments/<int:aid>/preview", methods=["GET"])
@login_required
def preview(plan_id, aid):
    _p, err = _guard_read(plan_id)
    if err:
        return err
    a = _get_att(plan_id, aid)
    if not a or not a.path or not os.path.exists(a.path):
        return jsonify({"ok": False, "error": "文件不存在"}), 404
    return send_file(a.path, as_attachment=False, download_name=a.filename)


@bp.route("/<int:plan_id>/attachments/<int:aid>/download", methods=["GET"])
@login_required
def download(plan_id, aid):
    _p, err = _guard_read(plan_id)
    if err:
        return err
    a = _get_att(plan_id, aid)
    if not a or not a.path or not os.path.exists(a.path):
        return jsonify({"ok": False, "error": "文件不存在"}), 404
    return send_file(a.path, as_attachment=True, download_name=a.filename)


@bp.route("/<int:plan_id>/attachments/<int:aid>", methods=["DELETE"])
@login_required
def delete_attachment(plan_id, aid):
    _plan, err = _load_and_guard(plan_id, "删除附件")
    if err:
        return err
    a = _get_att(plan_id, aid)
    if not a:
        return jsonify({"ok": False, "error": "附件不存在"}), 404
    # 从 WPS 导入的原始资料不删磁盘文件，只解除挂载——那是唯一一份存档
    if a.source == "upload" and a.path and os.path.exists(a.path):
        try:
            os.remove(a.path)
        except OSError:
            pass
    db.session.delete(a)
    db.session.commit()
    return jsonify({"ok": True, "message": "已删除"})

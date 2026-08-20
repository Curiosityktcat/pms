# -*- coding: utf-8 -*-
"""项目管理器：按当前账号的数据范围汇聚项目进度，只读展示，不承接业务录入。"""
import calendar
import datetime
import io
import os
import re
import secrets
from functools import wraps

from flask import Blueprint, jsonify, request, send_file, session
from sqlalchemy import case

from models import db
from models.procurement_plan import NOT_PROCURED, ProcurementPlan
from models.procurement_round import ProcurementRound
from models.project_file import ProjectFile
from models.project import Project
from models.sys_config import SysConfig
from routes.utils import login_required
from services.dept_scope import (assert_can_view_project, current_dept_code,
                                 is_dept_role, scope_projects)
from services.pending_owner import attach_pending, pending_map
from services.permission import get_user_perms, is_admin_user
from services.project_progress import build_progress, stage_map

bp = Blueprint("project_monitor", __name__, url_prefix="/api/project-monitor")

PERM_KEY = "project-monitor"
OVERDUE_CONFIG_KEY = "project_monitor_overdue_days"
DEFAULT_OVERDUE_DAYS = 7
ALLOWED_ROLES = ("dept_demand", "dept_manage", "officer", "leader", "assistant", "admin")

STAGE_LABELS = {
    "establish": "立项",
    "demand_confirm": "需求确认（5.1）",
    "inquiry": "询/议价函",
    "doc_confirm": "采购文件确认（5.2）",
    "announce": "公告发布",
    "bid_open": "开标",
    "review": "评审",
    "round_failed": "本轮流标",
    "result": "采购结果确认（9）",
    "contract": "合同签订（10）",
    "archive": "归档（11）",
}
STAGE_ORDER = list(STAGE_LABELS)


def _monitor_required(fn):
    """菜单权限之外再按角色白名单收口，防止误把权限配给代理机构后接口随之放开。"""
    @wraps(fn)
    @login_required
    def wrapper(*args, **kwargs):
        role = session.get("role", "")
        username = session.get("user", "")
        if role not in ALLOWED_ROLES and not is_admin_user(username):
            return jsonify({"ok": False, "error": "无权使用项目管理器"}), 403
        if PERM_KEY not in get_user_perms(username, role):
            return jsonify({"ok": False, "error": "无权使用项目管理器"}), 403
        return fn(*args, **kwargs)
    return wrapper


def _overdue_days():
    """阈值只认 SysConfig；配置损坏时回落默认值，不能让整个只读看板因此打不开。"""
    row = db.session.get(SysConfig, OVERDUE_CONFIG_KEY)
    try:
        value = int((row.value if row else "") or DEFAULT_OVERDUE_DAYS)
    except (TypeError, ValueError):
        value = DEFAULT_OVERDUE_DAYS
    return max(value, 1)


def _int_arg(name, default, minimum=1, maximum=None):
    try:
        value = int(request.args.get(name, default))
    except (TypeError, ValueError):
        value = default
    value = max(value, minimum)
    return min(value, maximum) if maximum is not None else value


def _filters():
    return {
        "year": (request.args.get("year") or "").strip().replace("年", ""),
        "manage_dept": (request.args.get("manage_dept") or "").strip(),
        "demand_dept": (request.args.get("demand_dept") or "").strip(),
        "officer": (request.args.get("officer") or "").strip(),
        "method": (request.args.get("method") or "").strip(),
        "stage": (request.args.get("stage") or "").strip(),
        "overdue": (request.args.get("overdue") or "").strip(),
        # 归档项目默认不进看板（这是「在办」看板），但科室常有整年项目全部归档、
        # 打开一片空白的情况，所以留一档让人主动把历史项目调出来看进度。
        "archived": (request.args.get("archived") or "").strip(),
        "keyword": (request.args.get("keyword") or "").strip(),
    }


def _visible_stmt(filters=None):
    """角色范围和普通筛选全部落在 SQL；草稿、删除不算项目，已归档默认不进看板。"""
    stmt = db.select(Project).where(
        db.or_(Project.is_draft == 0, Project.is_draft.is_(None)),
        db.or_(Project.is_deleted == 0, Project.is_deleted.is_(None)),
    )
    if (filters or {}).get("archived") != "1":
        stmt = stmt.where(db.or_(Project.status != "已归档", Project.status.is_(None)))
    role = session.get("role", "")
    if is_dept_role(role):
        stmt = scope_projects(stmt)
    elif role == "officer":
        stmt = stmt.where(Project.officer == session.get("display_name", ""))

    f = filters or {}
    if f.get("year"):
        stmt = stmt.where(db.func.replace(Project.year, "年", "") == f["year"])
    for key in ("manage_dept", "demand_dept", "officer", "method"):
        if f.get(key):
            stmt = stmt.where(getattr(Project, key) == f[key])
    if f.get("keyword"):
        like = f"%{f['keyword']}%"
        stmt = stmt.where(db.or_(Project.name.like(like), Project.number.like(like)))
    return stmt


def _display_stage(progress, pending):
    """展示阶段仍以 stage_map 为底，只用待办事件细分评审/合同这两个并行环节。"""
    stage = (progress or {}).get("current_stage") or ""
    event = (pending or {}).get("event", "")
    if event.startswith("review_") or event == "inquiry_review":
        return "review"
    if event.startswith("contract"):
        return "contract"
    if stage == "done":
        return "archive"
    return stage or "establish"


def _matching_context(filters):
    """先用 SQL 取可见 ID，再调用唯一进度/派单服务；不加载全院项目实体做 Python 过滤。"""
    id_stmt = _visible_stmt(filters).with_only_columns(Project.id).order_by(None)
    ids = list(db.session.execute(id_stmt).scalars())
    stages = stage_map(ids)
    pending = pending_map(ids)
    threshold = _overdue_days()

    matched = []
    display_stages = {}
    for pid in ids:
        p_stage = stages.get(pid) or {}
        p_pending = pending.get(pid)
        display = _display_stage(p_stage, p_pending)
        display_stages[pid] = display
        if filters.get("stage") and display != filters["stage"]:
            continue
        waiting = (p_pending or {}).get("waiting_days")
        is_overdue = waiting is not None and waiting > threshold
        if filters.get("overdue") == "1" and not is_overdue:
            continue
        if filters.get("overdue") == "0" and is_overdue:
            continue
        matched.append(pid)

    # “最该催的排最前”必须按 pending_owner 的同一 waiting_days 口径，不能另算一套。
    matched.sort(key=lambda pid: (
        (pending.get(pid) or {}).get("waiting_days") is not None,
        (pending.get(pid) or {}).get("waiting_days") or 0,
        pid,
    ), reverse=True)
    return matched, stages, pending, display_stages, threshold


def _ordered_projects(filters, ids, offset=0, limit=None):
    if not ids:
        return []
    order = case({pid: pos for pos, pid in enumerate(ids)}, value=Project.id, else_=len(ids))
    stmt = _visible_stmt(filters).where(Project.id.in_(ids)).order_by(order)
    if offset:
        stmt = stmt.offset(offset)
    if limit is not None:
        stmt = stmt.limit(limit)
    return list(db.session.execute(stmt).scalars())


def _project_rows(projects, stages, display_stages, threshold):
    rows = []
    for p in projects:
        info = stages.get(p.id) or {}
        rows.append({
            "id": p.id,
            "name": p.name or "",
            "number": p.number or "",
            "manage_dept": p.manage_dept or "",
            "demand_dept": p.demand_dept or "",
            "officer": p.officer or "",
            "method": p.method or "",
            "amount": p.amount,
            "current_stage": display_stages.get(p.id, "establish"),
            "stage_label": STAGE_LABELS.get(display_stages.get(p.id, ""), "进行中"),
            "current_round": info.get("current_round") or p.round or 1,
            "updated_at": _norm_date(p.updated_at or p.created_at),
        })
    # 最终返回必须显式走 attach_pending，确保列表字段和其它业务页的接线方式一致。
    attach_pending(rows)
    for row in rows:
        pending = row.get("pending") or {}
        waiting = pending.get("waiting_days")
        row["overdue"] = waiting is not None and waiting > threshold
        row["last_action_at"] = _norm_date(pending.get("since") or row["updated_at"])
    return rows


@bp.route("/meta", methods=["GET"])
@_monitor_required
def meta():
    """筛选项只从当前账号可见范围取，避免下拉框本身泄露别人的项目。"""
    stmt = _visible_stmt()

    def distinct(column):
        q = stmt.with_only_columns(column).order_by(None).where(column.isnot(None), column != "").distinct()
        return sorted({str(v).strip() for v in db.session.execute(q).scalars() if str(v).strip()})

    role = session.get("role", "")
    years = {v.replace("年", "") for v in distinct(Project.year)}
    if is_dept_role(role):
        # 科室可能只有尚未立项的年度计划，年度下拉不能只从正式项目里取，否则该年无从选择。
        from services.dept import dept_names
        names = dept_names(current_dept_code())
        plan_years = db.session.execute(
            db.select(ProcurementPlan.year).where(
                db.or_(ProcurementPlan.dept.in_(names),
                       ProcurementPlan.demand_dept.in_(names)),
                ProcurementPlan.year.isnot(None), ProcurementPlan.year != 0,
            ).distinct()
        ).scalars() if names else []
        years.update(str(value) for value in plan_years)
    return jsonify({"ok": True, "data": {
        "years": sorted(years, reverse=True),
        "manage_depts": distinct(Project.manage_dept),
        "demand_depts": distinct(Project.demand_dept),
        "officers": distinct(Project.officer),
        "methods": distinct(Project.method),
        "stages": [{"value": k, "label": STAGE_LABELS[k]} for k in STAGE_ORDER],
        "overdue_days": _overdue_days(),
        "show_officer_stats": role in ("leader", "admin") or is_admin_user(session.get("user", "")),
        "show_plans": is_dept_role(role),
    }})


@bp.route("/projects", methods=["GET"])
@_monitor_required
def projects():
    filters = _filters()
    page = _int_arg("page", 1)
    page_size = _int_arg("page_size", 20, maximum=100)
    ids, stages, _pending, display_stages, threshold = _matching_context(filters)
    page_rows = _ordered_projects(filters, ids, (page - 1) * page_size, page_size)
    return jsonify({"ok": True, "data": _project_rows(page_rows, stages, display_stages, threshold),
                    "total": len(ids), "page": page, "page_size": page_size})


@bp.route("/stats", methods=["GET"])
@_monitor_required
def stats():
    filters = _filters()
    ids, _stages, pending, display_stages, threshold = _matching_context(filters)
    month = datetime.date.today().strftime("%Y-%m")
    created = 0
    if ids:
        created = db.session.execute(
            db.select(db.func.count()).select_from(Project).where(
                Project.id.in_(ids), Project.created_at.like(f"{month}%"))
        ).scalar_one()

    by_stage = {}
    overdue = 0
    for pid in ids:
        stage = display_stages.get(pid, "establish")
        by_stage[stage] = by_stage.get(stage, 0) + 1
        waiting = (pending.get(pid) or {}).get("waiting_days")
        if waiting is not None and waiting > threshold:
            overdue += 1

    by_officer = []
    role = session.get("role", "")
    if ids and (role in ("leader", "admin") or is_admin_user(session.get("user", ""))):
        counts = db.session.execute(
            db.select(Project.officer, db.func.count()).where(Project.id.in_(ids))
            .group_by(Project.officer).order_by(db.func.count().desc())
        ).all()
        by_officer = [{"name": name or "（未指定）", "count": count} for name, count in counts]

    stages = [{"stage": key, "label": STAGE_LABELS.get(key, key), "count": by_stage[key]}
              for key in STAGE_ORDER if key in by_stage]
    return jsonify({"ok": True, "data": {
        "ongoing": len(ids), "new_this_month": created, "overdue": overdue,
        "by_stage": stages, "by_officer": by_officer, "overdue_days": threshold,
    }})


def _assert_visible_project(pid):
    if is_dept_role():
        assert_can_view_project(pid)
    p = db.session.get(Project, pid)
    if not p:
        return None, (jsonify({"ok": False, "error": "项目不存在"}), 404)
    if session.get("role") == "officer" and p.officer != session.get("display_name", ""):
        return None, (jsonify({"ok": False, "error": "无权查看该项目"}), 403)
    if p.is_draft or p.is_deleted:
        return None, (jsonify({"ok": False, "error": "项目不存在"}), 404)
    # 归档项目照样能看时间线：看板默认不列它，但点进来就是要看它当初怎么走完的。
    return p, None


def _norm_date(raw):
    """各环节存的日期格式不一（有 ISO、有「2026年6月24日」），看板上必须统一成 YYYY-MM-DD。"""
    text = str(raw or "").strip()
    if not text:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}", text):
        return text[:10]
    nums = re.findall(r"\d+", text)
    if len(nums) >= 3 and len(nums[0]) == 4:
        try:
            return datetime.date(int(nums[0]), int(nums[1]), int(nums[2])).isoformat()
        except ValueError:
            return text[:10]
    return text[:10]


def _milestone(key, label, source=None, done=False, at="", by=""):
    source = source or {}
    node_at = source.get("at") or at or ""
    if key == "contract" and not node_at:
        node_at = next((str(x.get("at") or "") for x in source.get("packages", []) if x.get("at")), "")
    return {"key": key, "label": label, "done": bool(source.get("done", done)),
            "at": _norm_date(node_at), "by": source.get("by") or by or ""}


@bp.route("/projects/<int:pid>/timeline", methods=["GET"])
@_monitor_required
def timeline(pid):
    """详情先做对象级可见性校验，再把既有进度服务整理成固定九节点横向时间线。"""
    p, err = _assert_visible_project(pid)
    if err:
        return err
    progress = build_progress(p)
    round_models = {r.round_number or 1: r for r in db.session.execute(
        db.select(ProcurementRound).filter_by(project_id=pid)
    ).scalars()}
    rounds = []
    # 部分历史项目没有轮次记录，也必须至少展示“已立项、后续未完成”，不能展开后一片空白。
    progress_rounds = progress.get("rounds") or [{"round_number": p.round or 1,
                                                  "status": "", "nodes": []}]
    for rd in progress_rounds:
        rn = rd.get("round_number") or 1
        source = {n.get("key"): n for n in rd.get("nodes", [])}
        model = round_models.get(rn)
        review_source = source.get("review")
        if not review_source and model:
            review_source = {"done": model.review_status == "已确认",
                             "at": model.review_confirmed_at or "",
                             "by": model.review_confirmed_by or ""}
        is_last = rn == (progress.get("project", {}).get("current_round") or 1)
        rounds.append({"round_number": rn, "status": rd.get("status") or "", "nodes": [
            _milestone("establish", "立项", done=True, at=p.created_at, by=p.created_by),
            _milestone("demand_confirm", "需求确认（5.1）", source.get("demand_confirm")),
            _milestone("doc_confirm", "采购文件确认（5.2）", source.get("doc_confirm")),
            _milestone("announce", "公告发布", source.get("announce")),
            _milestone("bid_open", "开标", source.get("bid_open")),
            _milestone("review", "评审", review_source),
            _milestone("result", "采购结果确认（9）", source.get("result")),
            _milestone("contract", "合同签订（10）", source.get("contract")),
            _milestone("archive", "归档（11）", done=is_last and p.status == "已归档",
                       at=p.updated_at if is_last and p.status == "已归档" else ""),
        ]})
    data = dict(progress)
    data["rounds"] = rounds
    return jsonify({"ok": True, "data": data})


# ══════════════════════════════════════════════════════════════════
# 项目资料：点开就能看这个项目归档文件夹里的东西
#
# 用户在《黄新博回应-WPS小团队搬入PMS方案》里写的：
#   「项目资料（可以点击后自动调取归档文件夹内的相关文件，还可以查看上传和删除，
#     上传还可以通过拖拽文件直接操作）」
# 复用 services.archive_print.list_archive_tree —— 归档那边早就把各模块的文件
# 按文件夹整理好了，这里不另收一套，否则两处口径会漂。
#
# 注意：这**不是**把「11. 归档」模块开放给科室。归档是流程的最后一步（要审要签），
# 这里只是让人看得到项目已经有哪些材料，是只读看板的一部分。
# ══════════════════════════════════════════════════════════════════

# 谁能往项目资料里传文件、删文件。科室和监督是只读的——
# 用户 2026-08-18 已把科室收成「除需求编制外只读」，上传删除属于写操作。
FILE_WRITE_ROLES = ("officer", "assistant", "leader", "admin")
# 补传的材料放这里，与各业务模块自己的目录分开，删起来不会误伤业务文件
MONITOR_UPLOAD_DIR = os.path.join(
    os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")),
    "uploads", "project_files")
ALLOWED_EXT = {".pdf", ".doc", ".docx", ".wps", ".xls", ".xlsx", ".ppt", ".pptx",
               ".jpg", ".jpeg", ".png", ".gif", ".zip", ".rar", ".txt", ".csv"}
MAX_UPLOAD_MB = 100


def _can_write_files():
    role = session.get("role", "")
    return role in FILE_WRITE_ROLES or is_admin_user(session.get("user", ""))


@bp.route("/projects/<int:pid>/files", methods=["GET"])
@_monitor_required
def project_files(pid):
    """这个项目的全部资料，按文件夹分组。"""
    p, err = _assert_visible_project(pid)
    if err:
        return err
    from services.archive_print import list_archive_tree
    try:
        folders = list_archive_tree(p)
    except Exception as e:                                   # noqa: BLE001
        # 某个环节的文件坏了不能让整个面板打不开
        folders = []
        print(f"[项目资料] 项目 {pid} 归档树读取失败：{e}", flush=True)

    extra_groups = {}
    for row in db.session.execute(
            db.select(ProjectFile).filter_by(project_id=pid)
            .order_by(ProjectFile.id.desc())).scalars():
        # 历史资料保留了原目录层级，几千个文件塞进一个组没法看；
        # folder 为空的仍归到原来那个组名，界面上老文件位置不变。
        folder = (row.folder or "").strip()
        group_name = folder or "补充材料（在项目管理器里传的）"
        extra_groups.setdefault(group_name, []).append({
            "id": row.id,
            "name": row.original_name or row.saved_name or "",
            "size": row.size or 0,
            "uploaded_by": row.uploaded_by or "",
            "uploaded_at": _norm_date(row.uploaded_at),
            "url": f"/api/project-monitor/projects/{pid}/files/{row.id}?download=1",
            "preview_url": f"/api/project-monitor/projects/{pid}/files/{row.id}",
            "can_delete": _can_write_files(),
        })
    if extra_groups:
        # 按组名排序，免得展示顺序随文件 id 变来变去
        folders = list(folders) + [
            {"folder": name, "items": extra_groups[name]}
            for name in sorted(extra_groups)
        ]

    total = sum(len(f.get("items") or []) for f in folders)
    return jsonify({"ok": True, "data": folders, "total": total,
                    "can_upload": _can_write_files()})


@bp.route("/projects/<int:pid>/files", methods=["POST"])
@_monitor_required
def upload_project_file(pid):
    """补传材料（支持拖拽，前端一次可以丢多个文件进来）。"""
    if not _can_write_files():
        return jsonify({"ok": False, "error": "只有采购部可以上传项目资料"}), 403
    p, err = _assert_visible_project(pid)
    if err:
        return err
    files = request.files.getlist("file") or request.files.getlist("files")
    files = [f for f in files if f and f.filename]
    if not files:
        return jsonify({"ok": False, "error": "没有收到文件"}), 400

    os.makedirs(os.path.join(MONITOR_UPLOAD_DIR, str(pid)), exist_ok=True)
    saved, rejected = [], []
    for f in files:
        name = os.path.basename(f.filename)
        ext = os.path.splitext(name)[1].lower()
        if ext not in ALLOWED_EXT:
            rejected.append(f"{name}（不支持的格式 {ext or '无扩展名'}）")
            continue
        blob = f.read()
        if len(blob) > MAX_UPLOAD_MB * 1024 * 1024:
            rejected.append(f"{name}（超过 {MAX_UPLOAD_MB}MB）")
            continue
        # 存盘名带随机串，避免同名互相覆盖；原名单独存，界面上显示原名
        stamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        saved_name = f"{stamp}_{secrets.token_hex(4)}{ext}"
        with open(os.path.join(MONITOR_UPLOAD_DIR, str(pid), saved_name), "wb") as fh:
            fh.write(blob)
        row = ProjectFile(project_id=pid, original_name=name, saved_name=saved_name,
                          folder="", size=len(blob),
                          uploaded_by=session.get("display_name", ""),
                          uploaded_at=datetime.datetime.now().isoformat(timespec="seconds"))
        db.session.add(row)
        saved.append(name)
    db.session.commit()
    if not saved:
        return jsonify({"ok": False, "error": "；".join(rejected) or "没有文件被保存"}), 400
    msg = f"已上传 {len(saved)} 个文件"
    if rejected:
        msg += f"；{len(rejected)} 个未收：" + "；".join(rejected)
    return jsonify({"ok": True, "message": msg, "saved": saved, "rejected": rejected})


@bp.route("/projects/<int:pid>/files/<int:fid>", methods=["GET"])
@_monitor_required
def get_project_file(pid, fid):
    p, err = _assert_visible_project(pid)
    if err:
        return err
    row = db.session.get(ProjectFile, fid)
    if not row or row.project_id != pid:
        return jsonify({"ok": False, "error": "文件不存在"}), 404
    path = os.path.join(MONITOR_UPLOAD_DIR, str(pid), row.saved_name or "")
    if not os.path.exists(path):
        return jsonify({"ok": False, "error": "文件已丢失"}), 404
    if request.args.get("download") == "1":
        return send_file(path, as_attachment=True,
                         download_name=row.original_name or row.saved_name)
    from services.office_convert import send_preview
    return send_preview(path, row.original_name or row.saved_name)


@bp.route("/projects/<int:pid>/files/<int:fid>", methods=["DELETE"])
@_monitor_required
def delete_project_file(pid, fid):
    if not _can_write_files():
        return jsonify({"ok": False, "error": "只有采购部可以删除项目资料"}), 403
    p, err = _assert_visible_project(pid)
    if err:
        return err
    row = db.session.get(ProjectFile, fid)
    if not row or row.project_id != pid:
        return jsonify({"ok": False, "error": "文件不存在"}), 404
    path = os.path.join(MONITOR_UPLOAD_DIR, str(pid), row.saved_name or "")
    if path and os.path.exists(path):
        try:
            os.remove(path)
        except OSError as e:                                 # noqa: BLE001
            print(f"[项目资料] 删文件失败 {path}: {e}", flush=True)
    name = row.original_name
    db.session.delete(row)
    db.session.commit()
    return jsonify({"ok": True, "message": f"已删除「{name}」"})


def _deadline_date(raw, default_year=None):
    text = (raw or "").strip()
    if not text:
        return None
    try:
        return datetime.date.fromisoformat(text[:10])
    except ValueError:
        pass
    nums = [int(x) for x in re.findall(r"\d+", text)]
    try:
        if len(nums) >= 3:
            return datetime.date(nums[0], nums[1], nums[2])
        if len(nums) == 2:
            return datetime.date(nums[0], nums[1], calendar.monthrange(nums[0], nums[1])[1])
        if len(nums) == 1 and default_year:
            year, month = int(default_year), nums[0]
            return datetime.date(year, month, calendar.monthrange(year, month)[1])
    except (ValueError, IndexError):
        return None
    return None


@bp.route("/plans", methods=["GET"])
@_monitor_required
def plans():
    """科室计划页签严格按会话科室码收口；采购部角色没有“我科室”概念，不开放此接口。"""
    if not is_dept_role():
        return jsonify({"ok": False, "error": "当前账号没有科室年度计划"}), 403
    from services.dept import dept_names
    names = dept_names(current_dept_code())
    if not names:
        return jsonify({"ok": False, "error": "未绑定科室"}), 403
    stmt = db.select(ProcurementPlan).where(db.or_(
        ProcurementPlan.dept.in_(names), ProcurementPlan.demand_dept.in_(names)))
    year = (request.args.get("year") or "").strip()
    keyword = (request.args.get("keyword") or "").strip()
    if year:
        try:
            stmt = stmt.where(ProcurementPlan.year == int(year))
        except ValueError:
            return jsonify({"ok": False, "error": "年度格式不正确"}), 400
    if keyword:
        like = f"%{keyword}%"
        stmt = stmt.where(db.or_(ProcurementPlan.name.like(like),
                                 ProcurementPlan.plan_number.like(like)))
    total = db.session.execute(
        db.select(db.func.count()).select_from(stmt.order_by(None).subquery())
    ).scalar_one()
    page = _int_arg("page", 1)
    page_size = _int_arg("page_size", 20, maximum=100)
    rows = list(db.session.execute(stmt.order_by(ProcurementPlan.id)
                                   .offset((page - 1) * page_size).limit(page_size)).scalars())
    pids = [r.project_id for r in rows if r.project_id]
    project_stmt = scope_projects(db.select(Project).where(Project.id.in_(pids)))
    projects = {p.id: p for p in db.session.execute(project_stmt).scalars()} if pids else {}
    visible_pids = list(projects)
    stages = stage_map(visible_pids)
    pending = pending_map(visible_pids)
    today = datetime.date.today()
    out = []
    for row in rows:
        project = projects.get(row.project_id)
        closed = (row.status or "") in NOT_PROCURED
        deadline = _deadline_date(row.deadline, row.year)
        overdue = bool(not project and not closed and deadline and deadline < today)
        near = bool(not project and not closed and deadline and today <= deadline <= today + datetime.timedelta(days=30))
        p_stage = _display_stage(stages.get(row.project_id) or {}, pending.get(row.project_id)) if project else ""
        out.append({
            "id": row.id, "year": row.year or 0, "name": row.name or "",
            "dept": row.dept or "", "demand_dept": row.demand_dept or "",
            "budget": row.budget or 0, "method": row.method or "", "status": row.status or "",
            "plan_status": (row.status or "") if closed else ("已立项" if project else "未立项"),
            "deadline": deadline.isoformat() if deadline else "",
            "deadline_raw": row.deadline or "", "overdue": overdue, "deadline_near": near,
            "project": ({"id": project.id, "number": project.number or "", "name": project.name or "",
                         "current_stage": p_stage, "stage_label": STAGE_LABELS.get(p_stage, "进行中"),
                         "pending": pending.get(project.id)} if project else None),
        })
    return jsonify({"ok": True, "data": out, "total": total,
                    "page": page, "page_size": page_size})


@bp.route("/export", methods=["GET"])
@_monitor_required
def export_projects():
    """导出和列表复用同一 filters/matching_context，不能出现页面与会议材料口径不同。"""
    filters = _filters()
    ids, stages, _pending, display_stages, threshold = _matching_context(filters)
    rows = _project_rows(_ordered_projects(filters, ids), stages, display_stages, threshold)
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "项目进度"
    headers = ["项目名称", "项目编号", "归口科室", "需求科室", "经办人", "采购方式",
               "预算（元）", "当前阶段", "当前处理人", "停留天数", "是否超期", "轮次", "最近动作时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        pending = row.get("pending") or {}
        ws.append([row["name"], row["number"], row["manage_dept"], row["demand_dept"],
                   row["officer"], row["method"], row["amount"], row["stage_label"],
                   pending.get("owner_name", ""), pending.get("waiting_days"),
                   "是" if row["overdue"] else "否", row["current_round"], row["last_action_at"]])
    for idx, width in enumerate((34, 20, 14, 14, 10, 18, 14, 20, 16, 10, 10, 8, 14), 1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"项目进度_{datetime.date.today().isoformat()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

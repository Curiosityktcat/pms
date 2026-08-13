"""采购项目分发（slice ②）。

采购部助理把项目（来自 rd-web 或手动）分发给经办人，并按规则自动指派代理机构。
可见性：经办人只看分给自己的；采购部助理/负责人/管理员看全部；其他角色看不到。
"""
import os
import uuid
import datetime

from flask import Blueprint, request, jsonify, session, send_file, current_app

from werkzeug.utils import secure_filename

from models import db
from models.project_distribution import ProjectDistribution, ProjectDistributionAttachment, RdwebAccount
from models.agency import Agency
from models.sys_config import SysConfig
from routes.utils import login_required
from services.permission import is_admin_user
from services import upload_relay

_MANUAL_THROTTLE_SEC = 30 * 60  # 手动刷新：30 分钟一次

bp = Blueprint("project_distribution", __name__, url_prefix="/api/distributions")

UPLOAD_ROOT = os.path.abspath(os.path.join(
    os.path.dirname(__file__), "..", "..", "uploads", "project_distribution"))

_MANAGE_ROLES = ("assistant", "pd_assistant", "leader")

# 项目池是「经办人本人的池子」：里面的审签表和附件是用黄新博本人的 rd-web 账号
# 抓下来的医院内部资料，所以经办人里只有他本人可见（其余经办人一律挡掉），
# 另加采购部助理/负责人/管理员。口径同「投标文件审核」，可用环境变量扩展。
POOL_OFFICERS = [s.strip() for s in os.environ.get("POOL_OFFICERS", "黄新博,agent-hxb").split(",") if s.strip()]


def _now():
    return datetime.datetime.now().isoformat(timespec="seconds")


def _attach_dir(did):
    d = os.path.join(UPLOAD_ROOT, str(did))
    os.makedirs(d, exist_ok=True)
    return d


def _can_manage():
    return session.get("role") in _MANAGE_ROLES or is_admin_user(session.get("user", ""))


def _pool_user():
    """能进项目池的人：白名单经办人本人 + 采购部助理/负责人/管理员。"""
    return _can_manage() or session.get("user", "") in POOL_OFFICERS


def _agency_name(code):
    if not code:
        return ""
    a = db.session.execute(db.select(Agency).filter_by(code=code)).scalar_one_or_none()
    return a.name if a else code


def _attachments(did):
    return db.session.execute(
        db.select(ProjectDistributionAttachment)
        .filter_by(distribution_id=did)
        .order_by(ProjectDistributionAttachment.id)
    ).scalars().all()


def _enrich(d):
    data = d.to_dict()
    data["agency_name"] = _agency_name(d.agency_code)
    data["attachments"] = [a.to_dict() for a in _attachments(d.id)]
    return data


# ── 从 rd-web 抓取（手动，30 分钟节流）───────────────────────────
@bp.route("/scrape-rdweb", methods=["POST"])
@login_required
def scrape_rdweb():
    if not _can_manage():
        return jsonify({"ok": False, "error": "仅采购部助理可抓取"}), 403
    # 节流：距上次抓取不足 30 分钟则拒绝
    row = db.session.get(SysConfig, "rdweb_last_scrape_at")
    if row and (row.value or "").strip():
        try:
            last = datetime.datetime.fromisoformat(row.value)
            left = _MANUAL_THROTTLE_SEC - (datetime.datetime.now() - last).total_seconds()
            if left > 0:
                return jsonify({"ok": False, "error": f"刚抓过，请 {int(left // 60) + 1} 分钟后再试"}), 429
        except ValueError:
            pass
    now = _now()
    if row is None:
        db.session.add(SysConfig(key="rdweb_last_scrape_at", value=now, updated_at=now))
    else:
        row.value = now
        row.updated_at = now
    db.session.commit()
    from services.rdweb_scraper import run_async
    started = run_async(current_app._get_current_object())
    if not started:
        return jsonify({"ok": False, "error": "已有抓取任务在进行中"}), 409
    return jsonify({"ok": True, "message": "已开始从 rd-web 抓取，约 1 分钟后刷新查看"})


@bp.route("/scrape-status", methods=["GET"])
@login_required
def scrape_status():
    from services.rdweb_scraper import _state
    return jsonify({"ok": True, "data": {"running": _state["running"], "last_msg": _state["last_msg"]}})


# ── rd-web 办理动作（接收/驳回/撤回，Playwright RPA，异步）──────────────
# ⚠️ 接收/驳回会真实盖陈梦霞电子签名并推进 rd-web 流程，前端须二次确认后才调用。
@bp.route("/<int:did>/rdweb-action", methods=["POST"])
@login_required
def rdweb_action(did):
    if not _can_manage():
        return jsonify({"ok": False, "error": "仅采购部助理可办理"}), 403
    d = db.session.get(ProjectDistribution, did)
    if not d:
        return jsonify({"ok": False, "error": "记录不存在"}), 404
    if not (d.serial_no or "").strip():
        return jsonify({"ok": False, "error": "该记录无流水号，无法在 rd-web 办理"}), 400
    data = request.get_json(silent=True) or {}
    action = (data.get("action") or "").strip()
    officer = (data.get("officer") or "").strip()
    opinion = (data.get("opinion") or "").strip()
    if action not in ("accept", "reject", "withdraw"):
        return jsonify({"ok": False, "error": "action 须为 accept/reject/withdraw"}), 400
    if action == "accept" and not officer:
        return jsonify({"ok": False, "error": "接收必须指定项目经办人"}), 400
    from services.rdweb_actions import run_async
    started = run_async(current_app._get_current_object(), action, d.serial_no, officer, opinion)
    if not started:
        return jsonify({"ok": False, "error": "已有 rd-web 办理任务在进行中，请稍候"}), 409
    return jsonify({"ok": True, "message": "已提交 rd-web 办理，约 1 分钟，请等待状态"})


@bp.route("/rdweb-action-status", methods=["GET"])
@login_required
def rdweb_action_status():
    from services.rdweb_actions import status
    s = status()
    return jsonify({"ok": True, "data": {"running": s["running"], "last_msg": s["last_msg"],
                                         "ok": s.get("ok"), "action": s.get("action"), "serial": s.get("serial")}})


# ── 列表（按可见性过滤）────────────────────────────────────────
@bp.route("", methods=["GET"])
@login_required
def list_distributions():
    name = session.get("display_name", "")
    # 项目池是黄新博本人用自己的 rd-web 账号抓来的医院内部资料：
    # 代理机构、其余经办人、监督人员一律不可见。
    if not _pool_user():
        return jsonify({"ok": False, "error": "项目池未对当前账号开放"}), 403
    q = db.select(ProjectDistribution).order_by(ProjectDistribution.id.desc())
    if not _can_manage():
        q = q.where(ProjectDistribution.officer == name)
    # 4.0 项目池（scope=pool）：只放经办人本人账号抓的（source='rd-经办人抓取'），
    # 通用抓取器/手动录入的属于「2. 采购项目分发」，不进这里。
    if request.args.get("scope") == "pool":
        q = q.where(ProjectDistribution.source == "rd-经办人抓取")
    rows = db.session.execute(q).scalars().all()
    return jsonify({"ok": True, "data": [_enrich(r) for r in rows]})


# ── 导出 Excel（按时间范围 / 采购方式筛选）──────────────────────
@bp.route("/export", methods=["GET"])
@login_required
def export_distributions():
    if not _can_manage():
        return jsonify({"ok": False, "error": "无权限"}), 403
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    # methods 多选（逗号分隔）；兼容旧的单个 method 参数
    methods = [m.strip() for m in (request.args.get("methods") or request.args.get("method") or "").split(",") if m.strip()]

    rows = db.session.execute(
        db.select(ProjectDistribution).order_by(ProjectDistribution.id.desc())
    ).scalars().all()

    def keep(d):
        dt = (d.created_at or "")[:10]
        if date_from and dt and dt < date_from:
            return False
        if date_to and dt and dt > date_to:
            return False
        if methods and d.method not in methods:
            return False
        return True

    rows = [d for d in rows if keep(d)]

    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "项目分发清单"
    headers = ["流水号", "项目名称", "归口管理科室", "需求科室", "预算金额(元)", "限价金额(元)",
               "采购组织形式", "采购方式", "项目编号", "项目内容", "经办人", "代理机构",
               "状态", "来源", "创建时间"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for d in rows:
        ws.append([
            d.serial_no, d.name, d.manage_dept, d.demand_dept, d.budget, d.price_limit,
            d.org_form, d.method, d.project_number, d.content, d.officer,
            _agency_name(d.agency_code), d.status, d.source, d.created_at,
        ])
    widths = [16, 32, 14, 12, 13, 13, 13, 13, 14, 40, 8, 16, 8, 8, 19]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    parts = ["项目分发清单"]
    if methods:
        parts.append("、".join(methods))
    if date_from or date_to:
        parts.append(f"{date_from or '起'}_{date_to or '今'}")
    fname = "_".join(parts) + ".xlsx"
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── 一键打印：把选中的 PDF 合并成一个 PDF（审签表排最前）──────────
@bp.route("/<int:did>/print", methods=["GET"])
@login_required
def print_merge(did):
    ids = {int(x) for x in (request.args.get("ids") or "").split(",") if x.strip().isdigit()}
    atts = [a for a in _attachments(did)
            if (not ids or a.id in ids) and (a.original_name or "").lower().endswith(".pdf")]
    atts.sort(key=lambda a: 0 if a.category == "审签表" else 1)
    if not atts:
        return jsonify({"ok": False, "error": "没有可打印的 PDF"}), 404
    import fitz  # PyMuPDF
    out = fitz.open()
    for a in atts:
        p = os.path.join(_attach_dir(did), a.saved_name)
        if os.path.exists(p):
            try:
                src = fitz.open(p)
                out.insert_pdf(src)
                src.close()
            except Exception:
                pass
    import io
    buf = io.BytesIO(out.tobytes())
    out.close()
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf",
                     download_name="一键打印.pdf", as_attachment=False)


# ── rd-web 账号维护（账号会变，给陈梦霞/经办人自维护）──────────────
def _sync_scrape_cfg(phone, password):
    """把"分发"账号同步到 SysConfig，供抓取程序使用。"""
    now = _now()
    for key, val in (("rdweb_loginuser", phone), ("rdweb_password", password)):
        row = db.session.get(SysConfig, key)
        if row is None:
            db.session.add(SysConfig(key=key, value=val, updated_at=now))
        else:
            row.value = val
            row.updated_at = now


@bp.route("/rdweb-accounts", methods=["GET"])
@login_required
def list_rdweb_accounts():
    rows = db.session.execute(db.select(RdwebAccount).order_by(RdwebAccount.id)).scalars().all()
    return jsonify({"ok": True, "data": [a.to_dict() for a in rows]})


@bp.route("/rdweb-accounts", methods=["POST"])
@login_required
def create_rdweb_account():
    data = request.get_json(force=True) or {}
    a = RdwebAccount(
        owner=(data.get("owner") or "").strip(),
        phone=(data.get("phone") or "").strip(),
        password=(data.get("password") or "").strip(),
        usage=(data.get("usage") or "执行").strip(),
        note=(data.get("note") or "").strip(),
        updated_at=_now(),
    )
    db.session.add(a)
    if a.usage == "分发" and a.phone:
        _sync_scrape_cfg(a.phone, a.password)
    db.session.commit()
    return jsonify({"ok": True, "data": a.to_dict()})


@bp.route("/rdweb-accounts/<int:aid>", methods=["PUT"])
@login_required
def update_rdweb_account(aid):
    a = db.session.get(RdwebAccount, aid)
    if not a:
        return jsonify({"ok": False, "error": "不存在"}), 404
    data = request.get_json(force=True) or {}
    for f in ("owner", "phone", "password", "usage", "note"):
        if f in data:
            setattr(a, f, (data[f] or "").strip())
    a.updated_at = _now()
    if a.usage == "分发" and a.phone:
        _sync_scrape_cfg(a.phone, a.password)
    db.session.commit()
    return jsonify({"ok": True, "data": a.to_dict()})


@bp.route("/rdweb-accounts/<int:aid>", methods=["DELETE"])
@login_required
def delete_rdweb_account(aid):
    a = db.session.get(RdwebAccount, aid)
    if a:
        db.session.delete(a)
        db.session.commit()
    return jsonify({"ok": True})


# ── 新建/分发（自动指派代理）──────────────────────────────────
@bp.route("", methods=["POST"])
@login_required
def create_distribution():
    if not _can_manage():
        return jsonify({"ok": False, "error": "仅采购部助理可分发项目"}), 403
    data = request.get_json(force=True) or {}
    method = (data.get("method") or "").strip()
    is_central = 1 if data.get("is_central") else 0
    officer = (data.get("officer") or "").strip()

    # 按规则自动指派代理机构（推进轮派指针）
    from services.agency_rotation import assign_agency
    ag = assign_agency(method, is_central=bool(is_central))
    agency_code = ag.code if ag else ""

    now = _now()
    d = ProjectDistribution(
        serial_no=(data.get("serial_no") or "").strip(),
        originator=(data.get("originator") or "").strip(),
        name=(data.get("name") or "").strip(),
        content=data.get("content") or "",
        manage_dept=(data.get("manage_dept") or "").strip(),
        demand_dept=(data.get("demand_dept") or "").strip(),
        budget=data.get("budget"),
        price_limit=data.get("price_limit"),
        method=method,
        org_form=(data.get("org_form") or "").strip(),
        project_number=(data.get("project_number") or "").strip(),
        is_central=is_central,
        officer=officer,
        agency_code=agency_code,
        source=(data.get("source") or "手动"),
        status="已分发" if officer else "待分发",
        created_by=session.get("display_name", ""),
        created_at=now,
        updated_at=now,
    )
    db.session.add(d)
    db.session.commit()
    return jsonify({"ok": True, "data": _enrich(d)})


# ── 更新（不自动重派代理；改派走专用接口）────────────────────
@bp.route("/<int:did>", methods=["PUT"])
@login_required
def update_distribution(did):
    if not _can_manage():
        return jsonify({"ok": False, "error": "无权限"}), 403
    d = db.session.get(ProjectDistribution, did)
    if not d:
        return jsonify({"ok": False, "error": "不存在"}), 404
    data = request.get_json(force=True) or {}
    for f in ("serial_no", "originator", "name", "content", "manage_dept", "demand_dept",
              "budget", "price_limit", "method", "org_form", "project_number", "officer"):
        if f in data:
            setattr(d, f, data[f])
    if "is_central" in data:
        d.is_central = 1 if data["is_central"] else 0
    if d.officer and d.status == "待分发":
        d.status = "已分发"
    d.updated_at = _now()
    db.session.commit()
    return jsonify({"ok": True, "data": _enrich(d)})


# ── 手动改派代理（重新走轮派或指定）──────────────────────────
@bp.route("/<int:did>/reassign-agency", methods=["POST"])
@login_required
def reassign_agency(did):
    if not _can_manage():
        return jsonify({"ok": False, "error": "无权限"}), 403
    d = db.session.get(ProjectDistribution, did)
    if not d:
        return jsonify({"ok": False, "error": "不存在"}), 404
    data = request.get_json(force=True) or {}
    code = (data.get("agency_code") or "").strip()
    if code:
        d.agency_code = code  # 指定某家（不推进轮派指针）
    else:
        from services.agency_rotation import assign_agency
        ag = assign_agency(d.method, is_central=bool(d.is_central))
        d.agency_code = ag.code if ag else ""
    d.updated_at = _now()
    db.session.commit()
    return jsonify({"ok": True, "data": _enrich(d)})


# ── 经办人对账：关联已立项项目 / 标记不立项 ────────────────────
# 从项目池点「立项」会自动回填 project_id+已立项；但经办人若走 4.1 项目立项
# 另行建项（立项时项目常被改名，无法按名称自动匹配），池里就一直挂着「待分发」。
@bp.route("/<int:did>/link-project", methods=["POST"])
@login_required
def link_project(did):
    d = db.session.get(ProjectDistribution, did)
    if not d:
        return jsonify({"ok": False, "error": "不存在"}), 404
    if not _visible(d):
        return jsonify({"ok": False, "error": "无权限"}), 403
    data = request.get_json(silent=True) or {}

    if data.get("cancel"):
        d.project_id = None
        d.status = "已取消"
        d.updated_at = _now()
        db.session.commit()
        return jsonify({"ok": True, "message": "已标记为不立项", "data": _enrich(d)})

    pid = data.get("project_id")
    if not pid:
        return jsonify({"ok": False, "error": "未选择项目"}), 400
    from models.project import Project
    p = db.session.get(Project, int(pid))
    if not p or p.is_deleted:
        return jsonify({"ok": False, "error": "项目不存在"}), 404
    # 经办人只能关联自己的项目；助理/负责人不限
    if session.get("role") == "officer" and p.officer != session.get("display_name", ""):
        return jsonify({"ok": False, "error": "只能关联自己经办的项目"}), 403
    other = db.session.execute(
        db.select(ProjectDistribution)
        .where(ProjectDistribution.project_id == p.id)
        .where(ProjectDistribution.id != d.id)
    ).scalars().first()
    if other:
        return jsonify({"ok": False, "error": f"该项目已关联池内条目「{other.name}」"}), 400
    d.project_id = p.id
    d.status = "已立项"
    d.updated_at = _now()
    db.session.commit()
    return jsonify({"ok": True, "message": f"已关联项目 {p.number or p.name}", "data": _enrich(d)})


@bp.route("/<int:did>/unlink-project", methods=["POST"])
@login_required
def unlink_project(did):
    """撤销关联/取消标记，退回「已分发」。"""
    d = db.session.get(ProjectDistribution, did)
    if not d:
        return jsonify({"ok": False, "error": "不存在"}), 404
    if not _visible(d):
        return jsonify({"ok": False, "error": "无权限"}), 403
    d.project_id = None
    d.status = "已分发" if d.officer else "待分发"
    d.updated_at = _now()
    db.session.commit()
    return jsonify({"ok": True, "message": "已撤销关联", "data": _enrich(d)})


# ── 删除 ──────────────────────────────────────────────────────
@bp.route("/<int:did>", methods=["DELETE"])
@login_required
def delete_distribution(did):
    if not _can_manage():
        return jsonify({"ok": False, "error": "无权限"}), 403
    d = db.session.get(ProjectDistribution, did)
    if not d:
        return jsonify({"ok": False, "error": "不存在"}), 404
    for a in _attachments(did):
        try:
            os.remove(os.path.join(_attach_dir(did), a.saved_name))
        except OSError:
            pass
        db.session.delete(a)
    db.session.delete(d)
    db.session.commit()
    return jsonify({"ok": True})


# ── 附件：上传 / 预览 / 下载 / 删除 ───────────────────────────
def _visible(d):
    if _can_manage():
        return True
    # 白名单经办人本人，且只看分给自己的
    return _pool_user() and d.officer == session.get("display_name", "")


@bp.route("/<int:did>/attachments", methods=["POST"])
@login_required
def upload_attachment(did):
    if not _can_manage():
        return jsonify({"ok": False, "error": "无权限"}), 403
    d = db.session.get(ProjectDistribution, did)
    if not d:
        return jsonify({"ok": False, "error": "不存在"}), 404
    f = request.files.get("file") or upload_relay.staged_file()  # 公网大文件走 OSS 中转（见 services/upload_relay.py）
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "未选择文件"}), 400
    ext = os.path.splitext(secure_filename(f.filename))[1]
    saved = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(_attach_dir(did), saved)
    f.save(path)
    att = ProjectDistributionAttachment(
        distribution_id=did, original_name=f.filename, saved_name=saved,
        file_size=os.path.getsize(path), mime_type=f.mimetype or "",
        uploaded_by=session.get("display_name", ""), uploaded_at=_now())
    db.session.add(att)
    db.session.commit()
    return jsonify({"ok": True, "data": att.to_dict()})


def _get_att(did, aid):
    d = db.session.get(ProjectDistribution, did)
    att = db.session.get(ProjectDistributionAttachment, aid)
    if not d or not att or att.distribution_id != did or not _visible(d):
        return None, None
    return d, att


@bp.route("/<int:did>/attachments/<int:aid>/preview", methods=["GET"])
@login_required
def preview_attachment(did, aid):
    d, att = _get_att(did, aid)
    if not att:
        return jsonify({"ok": False, "error": "无权限或不存在"}), 404
    from services.office_convert import send_preview
    return send_preview(os.path.join(_attach_dir(did), att.saved_name), att.original_name)


@bp.route("/<int:did>/attachments/<int:aid>/download", methods=["GET"])
@login_required
def download_attachment(did, aid):
    d, att = _get_att(did, aid)
    if not att:
        return jsonify({"ok": False, "error": "无权限或不存在"}), 404
    return send_file(os.path.join(_attach_dir(did), att.saved_name),
                     download_name=att.original_name, as_attachment=True)


@bp.route("/<int:did>/attachments/<int:aid>", methods=["DELETE"])
@login_required
def delete_attachment(did, aid):
    if not _can_manage():
        return jsonify({"ok": False, "error": "无权限"}), 403
    d, att = _get_att(did, aid)
    if not att:
        return jsonify({"ok": False, "error": "不存在"}), 404
    try:
        os.remove(os.path.join(_attach_dir(did), att.saved_name))
    except OSError:
        pass
    db.session.delete(att)
    db.session.commit()
    return jsonify({"ok": True})


# ── 抓取「分发给我(经办人)的采购需求审签表·待处理」→ 项目池（含附件，rdweb_mine）──
@bp.route("/scrape-mine", methods=["POST"])
@login_required
def scrape_mine():
    # 抓的是绑定在 rdweb_mine 里的那个本人账号（黄新博），别的经办人抓来的不是自己的东西
    if not _pool_user():
        return jsonify({"ok": False, "error": "项目池未对当前账号开放"}), 403
    from services.rdweb_mine import run_async
    started = run_async(current_app._get_current_object())
    if not started:
        return jsonify({"ok": False, "error": "已有抓取任务在进行中，请稍候"}), 409
    return jsonify({"ok": True, "message": "已开始抓取分发给我的待处理审签表(含附件)，约1-2分钟后刷新项目池"})


@bp.route("/scrape-mine-status", methods=["GET"])
@login_required
def scrape_mine_status():
    from services.rdweb_mine import status
    return jsonify({"ok": True, "data": status()})

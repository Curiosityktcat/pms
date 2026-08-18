"""「含已归档」这一档的验收：默认仍只看在办，打开后能看到历史项目，且不破坏隔离。"""
import os, sys, secrets
sys.path.insert(0, ".")
os.environ["PMS_DB_PATH"] = "/home/huangxb/pms/pms.test.db"
os.environ["PMS_CAPTCHA_ON"] = "0"
PW = "Monitor!2026"
ok, bad = 0, []


def check(t, c, e=""):
    global ok
    if c: ok += 1; print(f"OK   {t} {e}")
    else: bad.append(t); print(f"FAIL {t} {e}")


from app import create_app
from models import db
from models.user import User
from models.project import Project
from services.auth import hash_pw
from services.dept import dept_names

app = create_app(); app.app_context().push()
assert "pms.test.db" in app.config["SQLALCHEMY_DATABASE_URI"], "保险丝：不是测试库"


def acct(code):
    u = db.session.execute(db.select(User).filter_by(dept_code=code).where(
        User.role.in_(("dept", "dept_manage", "dept_demand")))).scalars().first()
    salt = secrets.token_hex(16); u.salt, u.pw_hash, u.active = salt, hash_pw(PW, salt), 1
    db.session.commit(); return u.username


def login(un):
    c = app.test_client()
    c.post("/api/auth/login", json={"username": un, "password": PW})
    return c


def ids(c, **q):
    out, page = set(), 1
    while True:
        r = c.get("/api/project-monitor/projects", query_string={**q, "page": page, "page_size": 100})
        if r.status_code != 200: return None
        b = r.get_json() or {}; rows = b.get("data") or []
        out |= {x["id"] for x in rows}
        if not rows or len(out) >= (b.get("total") or 0): return out
        page += 1


zw = login(acct("ZWK"))       # 总务科：4 个项目全部已归档
sb = login(acct("SBK"))       # 医学装备部

check("① 默认仍只看在办（总务科为空）", ids(zw) == set(), f"{len(ids(zw))} 个")
arch = ids(zw, archived="1")
check("② 打开「含已归档」后看得到历史项目", len(arch) == 4, f"{len(arch)} 个")
names = dept_names("ZWK")
rows = db.session.execute(db.select(Project.id, Project.status, Project.manage_dept,
                                    Project.demand_dept).where(Project.id.in_(arch))).all()
check("② 调出来的确实是本科室的", all(md in names or dd in names for _, _, md, dd in rows))
check("② 且确实是已归档的", all(st == "已归档" for _, st, _, _ in rows), f"{[r[1] for r in rows]}")

# 隔离不能因为这一档松掉
a_all, b_all = ids(zw, archived="1"), ids(sb, archived="1")
check("③ 含归档时两科室仍无交集", not (a_all & b_all), f"{len(a_all)} / {len(b_all)}，交集 {len(a_all & b_all)}")
victim = sorted(b_all - a_all)[0]
r = zw.get(f"/api/project-monitor/projects/{victim}/timeline")
check("③ 含归档时越权仍 403", r.status_code == 403, f"HTTP {r.status_code}")

# 归档项目的时间线要能打开（本科室自己的）
pid = sorted(arch)[0]
r = zw.get(f"/api/project-monitor/projects/{pid}/timeline")
check("④ 已归档项目的时间线能打开", r.status_code == 200, f"HTTP {r.status_code}")
nodes = ((r.get_json() or {}).get("data") or {}).get("rounds", [{}])[0].get("nodes") or []
last = nodes[-1] if nodes else {}
check("④ 归档节点标为已完成", last.get("key") == "archive" and last.get("done"),
      f"{last.get('label')} done={last.get('done')} at={last.get('at')}")

# 卡片与列表在这一档下也要一致
r1 = zw.get("/api/project-monitor/stats", query_string={"archived": "1"})
r2 = zw.get("/api/project-monitor/projects", query_string={"archived": "1", "page": 1, "page_size": 1})
check("⑤ 含归档时卡片与列表口径一致",
      (r1.get_json()["data"]["ongoing"]) == (r2.get_json()["total"]),
      f'{r1.get_json()["data"]["ongoing"]} vs {r2.get_json()["total"]}')

# 导出跟随该筛选
r = zw.get("/api/project-monitor/export", query_string={"archived": "1"})
import io as _io
from openpyxl import load_workbook
n = load_workbook(_io.BytesIO(r.data)).active.max_row - 1
check("⑥ 导出跟随「含已归档」", n == 4, f"{n} 行")

print(f"\n通过 {ok} 项" + (f"，失败 {len(bad)}：{bad}" if bad else "，全部通过"))
sys.exit(1 if bad else 0)

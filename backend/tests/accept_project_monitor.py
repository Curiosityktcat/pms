"""项目管理器验收（只打测试实例 1574 / pms.test.db）。

对着 PROJECT_MONITOR_TASK.md 第六节 8 条逐条打。
重点：四类账号的可见范围、卡天数倒序、卡片与列表口径一致、越权 403、代理机构看不到。
"""
import os
import sys
import secrets

sys.path.insert(0, ".")
os.environ["PMS_DB_PATH"] = "/home/huangxb/pms/pms.test.db"
# 进程内直打 test_client：不碰公网测试实例(1574)的滑块开关，也不留下弱配置。
os.environ["PMS_CAPTCHA_ON"] = "0"

PW = "Monitor!2026"
ok_count, bad = 0, []


def check(tag, cond, extra=""):
    global ok_count
    if cond:
        ok_count += 1
        print(f"OK   {tag} {extra}")
    else:
        bad.append(tag)
        print(f"FAIL {tag} {extra}")


from app import create_app
from models import db
from models.user import User
from models.dept import Dept
from models.project import Project
from services.auth import hash_pw

app = create_app()
DEPT_ROLES = ("dept", "dept_manage", "dept_demand")


def setpw(u):
    salt = secrets.token_hex(16)
    u.salt, u.pw_hash, u.active = salt, hash_pw(PW, salt), 1
    if hasattr(u, "must_change_pw"):
        u.must_change_pw = 0
    return u


def ensure_dept_account(code):
    d = db.session.execute(db.select(Dept).filter_by(code=code)).scalar_one()
    u = db.session.execute(db.select(User).filter_by(dept_code=code).where(
        User.role.in_(DEPT_ROLES))).scalars().first()
    if u is None:
        # 「dept」是已废弃的老角色（2026-08-18 全部迁走了），新建必须按科室类型
        # 给正确角色，否则建出来的账号连项目管理器的角色白名单都过不去。
        from routes.user_admin_api import _dept_role
        u = User(username=d.name, display_name=d.name, role=_dept_role(d), dept_code=code,
                 active=1, agency_code="", salt="", pw_hash="")
        db.session.add(u)
    setpw(u)
    return u.username, d.name


def ensure_role_account(username, display, role, dept_code=""):
    u = db.session.execute(db.select(User).filter_by(username=username)).scalar_one_or_none()
    if u is None:
        u = User(username=username, display_name=display, role=role, dept_code=dept_code,
                 active=1, agency_code="", salt="", pw_hash="")
        db.session.add(u)
    u.display_name, u.role = display, role
    setpw(u)
    return username


with app.app_context():
    uri = app.config["SQLALCHEMY_DATABASE_URI"]
    assert "pms.test.db" in uri, f"保险丝：不是测试库 {uri}"
    a_user, a_name = ensure_dept_account("SBK")     # 医学装备部（归口）
    b_user, b_name = ensure_dept_account("ZWK")     # 总务科
    # 经办人：挑一个库里真的经办过项目的人，否则范围测试没意义
    top_officer = db.session.execute(
        db.select(Project.officer, db.func.count()).where(
            Project.officer.isnot(None), Project.officer != "")
        .group_by(Project.officer).order_by(db.func.count().desc())
    ).first()
    officer_name = top_officer[0]
    officer_total = top_officer[1]
    off_user = ensure_role_account("acc_officer", officer_name, "officer")
    lead_user = ensure_role_account("acc_leader", "验收主任", "leader")
    agency_user = ensure_role_account("acc_agency", "验收代理", "agency")
    db.session.commit()
    print(f"账号：科室A={a_user}({a_name}) 科室B={b_user}({b_name}) "
          f"经办人={officer_name}(库里 {officer_total} 个) 主任={lead_user} 代理={agency_user}\n")


_ctx = app.app_context()
_ctx.push()


def login(username):
    c = app.test_client()
    r = c.post("/api/auth/login", json={"username": username, "password": PW})
    body = r.get_json() or {}
    return c, (body.get("user") or {})


def get(sess, path, **params):
    r = sess.get(path, query_string=params)
    try:
        return r.status_code, (r.get_json() or {})
    except Exception:
        return r.status_code, {}


sa, ua = login(a_user)
sb, ub = login(b_user)
so, uo = login(off_user)
sl, ul = login(lead_user)
sg, ug = login(agency_user)
check("① 五个账号都能登录", all([ua, ub, uo, ul, ug]),
      f"{ua.get('role')}/{ub.get('role')}/{uo.get('role')}/{ul.get('role')}/{ug.get('role')}")

# ── ① 可见范围 ──────────────────────────────────────────
def all_ids(sess):
    ids, page = set(), 1
    while True:
        sc, body = get(sess, "/api/project-monitor/projects", page=page, page_size=100)
        if sc != 200:
            return None, sc
        rows = body.get("data") or []
        ids |= {r["id"] for r in rows}
        if len(ids) >= (body.get("total") or 0) or not rows:
            return ids, 200
        page += 1


ia, ca = all_ids(sa)
ib, cb = all_ids(sb)
io_, co = all_ids(so)
il, cl = all_ids(sl)
check("① 科室/经办人/主任都能拉到列表", None not in (ia, ib, io_, il),
      f"A {len(ia or [])} / B {len(ib or [])} / 经办 {len(io_ or [])} / 主任 {len(il or [])}")
if ia is not None and ib is not None:
    check("① 两科室可见集合无交集", not (ia & ib), f"交集 {len(ia & ib)} 个")
if il is not None and ia is not None:
    check("① 主任看到的是全院（覆盖科室A）", ia <= il and len(il) >= len(ia), f"{len(il)} ⊇ {len(ia)}")

if True:
    if io_ is not None:
        rows = db.session.execute(db.select(Project.id, Project.officer).where(
            Project.id.in_(io_))).all() if io_ else []
        wrong = [pid for pid, off in rows if (off or "") != officer_name]
        check("① 经办人只看到自己经办的", not wrong, f"越界 {len(wrong)} 个")

# ── ⑧ 代理机构 ──────────────────────────────────────────
sc_ag, _ = get(sg, "/api/project-monitor/projects")
check("⑧ 代理机构调接口被拒", sc_ag == 403, f"HTTP {sc_ag}")
check("⑧ 代理机构没有该菜单权限", "project-monitor" not in set(ug.get("perms", [])))
for who, u in (("科室A", ua), ("经办人", uo), ("主任", ul)):
    check(f"⑧ {who}有该菜单权限", "project-monitor" in set(u.get("perms", [])))

# ── ② 默认按卡的天数倒序 ─────────────────────────────────
sc, body = get(sl, "/api/project-monitor/projects", page=1, page_size=20)
rows = body.get("data") or []
days = [((r.get("pending") or {}).get("waiting_days")) for r in rows]
have = [d for d in days if d is not None]
check("② 列表默认按卡的天数倒序", have == sorted(have, reverse=True),
      f"前 5: {days[:5]}")
check("② 最上面那条确实卡最久", bool(have) and days[0] == max(have),
      f"首条 {days[0] if days else '无数据'} 天")

# ── ③ 时间线 ───────────────────────────────────────────
if rows:
    pid = rows[0]["id"]
    sc, body = get(sl, f"/api/project-monitor/projects/{pid}/timeline")
    rds = (body.get("data") or {}).get("rounds") or []
    check("③ 时间线能取到", sc == 200 and bool(rds), f"HTTP {sc}，{len(rds)} 轮")
    if rds:
        nodes = rds[0].get("nodes") or []
        check("③ 固定九节点", len(nodes) == 9, f"{len(nodes)} 个")
        done = [n for n in nodes if n.get("done")]
        check("③ 已完成节点带日期", all(n.get("at") for n in done), f"{len(done)} 个已完成")
        check("③ 立项节点必定完成", bool(nodes) and nodes[0]["done"] and nodes[0]["at"])
    # 找一个多轮项目
    if True:
        multi = db.session.execute(db.select(Project.id).where(Project.round > 1)).scalars().first()
    if multi:
        sc, body = get(sl, f"/api/project-monitor/projects/{multi}/timeline")
        n = len((body.get("data") or {}).get("rounds") or [])
        check("③ 多轮项目能切轮次", sc == 200 and n >= 1, f"{n} 轮")

# ── ④ 卡片与列表同一口径 ─────────────────────────────────
for label, sess in (("主任", sl), ("科室A", sa)):
    sc1, st = get(sess, "/api/project-monitor/stats")
    sc2, li = get(sess, "/api/project-monitor/projects", page=1, page_size=1)
    a1 = (st.get("data") or {}).get("ongoing")
    a2 = li.get("total")
    check(f"④ {label} 卡片在办数=列表总数", a1 == a2, f"{a1} vs {a2}")
    stg = (st.get("data") or {}).get("by_stage") or []
    check(f"④ {label} 阶段分布合计=在办数", sum(x["count"] for x in stg) == a1,
          f"{sum(x['count'] for x in stg)} vs {a1}")
# 带筛选也要一致
sc1, st = get(sl, "/api/project-monitor/stats", overdue="1")
sc2, li = get(sl, "/api/project-monitor/projects", overdue="1", page=1, page_size=1)
check("④ 筛「仅超期」后卡片与列表仍一致",
      (st.get("data") or {}).get("ongoing") == li.get("total"),
      f"{(st.get('data') or {}).get('ongoing')} vs {li.get('total')}")
check("④ 超期数与仅超期筛选条数一致",
      (st.get("data") or {}).get("overdue") == li.get("total"),
      f"overdue={(st.get('data') or {}).get('overdue')} 筛出 {li.get('total')}")

# ── ⑤ 我科室的年度计划 ───────────────────────────────────
sc, body = get(sa, "/api/project-monitor/plans", page=1, page_size=100)
plans = body.get("data") or []
check("⑤ 科室能看到年度计划页签", sc == 200, f"HTTP {sc}，{body.get('total')} 条")
linked = [p for p in plans if p.get("project")]
check("⑤ 已立项计划显示对应项目进度",
      not linked or all(p["project"].get("stage_label") for p in linked),
      f"{len(linked)} 条已关联项目")
od = [p for p in plans if p.get("overdue")]
check("⑤ 过期未立项能标出", all(not p.get("project") for p in od), f"{len(od)} 条标红")
sc_lp, _ = get(sl, "/api/project-monitor/plans")
check("⑤ 采购部角色没有「我科室计划」", sc_lp == 403, f"HTTP {sc_lp}")

# ── ⑥ 越权 403 ─────────────────────────────────────────
if ib:
    victim = sorted(ib - (ia or set()))[0]
    sc, _ = get(sa, f"/api/project-monitor/projects/{victim}/timeline")
    check("⑥ 科室改URL取他科室时间线被挡", sc in (403, 404), f"HTTP {sc}")
if il and io_:
    other = sorted(il - io_)
    if other:
        sc, _ = get(so, f"/api/project-monitor/projects/{other[0]}/timeline")
        check("⑥ 经办人取别人项目时间线被挡", sc in (403, 404), f"HTTP {sc}")

# ── ⑦ 导出与筛选一致 ────────────────────────────────────
r = sa.get("/api/project-monitor/export")
check("⑦ 导出 Excel 能下载", r.status_code == 200 and len(r.data) > 2000,
      f"HTTP {r.status_code}，{len(r.data)} 字节")
if r.status_code == 200:
    import io as _io
    from openpyxl import load_workbook
    ws = load_workbook(_io.BytesIO(r.data)).active
    n = ws.max_row - 1
    sc, li = get(sa, "/api/project-monitor/projects", page=1, page_size=1)
    check("⑦ 导出条数=当前筛选条数", n == li.get("total"), f"{n} vs {li.get('total')}")

# ── 只读：非 GET 一律不通 ────────────────────────────────
r = sa.post("/api/project-monitor/projects", json={})
check("⑨ 本模块无写接口", r.status_code in (403, 404, 405), f"HTTP {r.status_code}")

print(f"\n通过 {ok_count} 项" + (f"，失败 {len(bad)} 项：{bad}" if bad else "，全部通过"))
sys.exit(1 if bad else 0)

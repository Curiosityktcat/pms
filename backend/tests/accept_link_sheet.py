"""Excel 对照表一来一回的验收（只打测试库）。

用户 2026-08-18：「你可以做一个项目清单的 excel 表，然后我把小团队项目名称和
项目编号手动关联一下」。重点不是能不能导出，而是**填错了系统拦不拦得住**。
"""
import io as _io
import os, sys, secrets
sys.path.insert(0, ".")
os.environ["PMS_CAPTCHA_ON"] = "0"
os.environ["PMS_DB_PATH"] = "/home/huangxb/pms/pms.test.db"
PW = "Sheet!2026"
ok, bad = 0, []


def check(t, c, e=""):
    global ok
    if c: ok += 1; print(f"OK   {t} {e}")
    else: bad.append(t); print(f"FAIL {t} {e}")


from openpyxl import load_workbook, Workbook
from app import create_app
from models import db
from models.user import User
from models.dept import Dept
from models.project import Project
from models.procurement_plan import ProcurementPlan
from services.auth import hash_pw

app = create_app(); app.app_context().push()
assert "pms.test.db" in app.config["SQLALCHEMY_DATABASE_URI"], "保险丝：不是测试库"
DEPT_ROLES = ("dept", "dept_manage", "dept_demand")


def acct(code):
    d = db.session.execute(db.select(Dept).filter_by(code=code)).scalar_one()
    u = db.session.execute(db.select(User).filter_by(dept_code=code).where(
        User.role.in_(DEPT_ROLES))).scalars().first()
    if u is None:
        from routes.user_admin_api import _dept_role
        u = User(username=d.name, display_name=d.name, role=_dept_role(d),
                 dept_code=code, active=1, agency_code="", salt="", pw_hash="")
        db.session.add(u)
    salt = secrets.token_hex(16); u.salt, u.pw_hash, u.active = salt, hash_pw(PW, salt), 1
    if hasattr(u, "must_change_pw"): u.must_change_pw = 0
    db.session.commit()
    return u.username


def login(un):
    c = app.test_client()
    c.post("/api/auth/login", json={"username": un, "password": PW})
    return c


sa = login(acct("SBK"))     # 医学装备部
sb = login(acct("ZWK"))     # 总务科

# 记下改动前的关联，跑完回滚
before = {r.id: r.project_id for r in db.session.execute(
    db.select(ProcurementPlan)).scalars()}

# ═══ 一、导出 ═══════════════════════════════════════════════════
print("── 一、导出对照表 ──")
r = sa.get("/api/procurement-plans/link-sheet")
check("① 能导出", r.status_code == 200 and len(r.data) > 3000,
      f"HTTP {r.status_code}，{len(r.data)} 字节")
wb = load_workbook(_io.BytesIO(r.data))
check("① 三张工作表齐全",
      set(wb.sheetnames) == {"计划对照表", "项目清单（查编号用）", "怎么填"},
      f"{wb.sheetnames}")
ws = wb["计划对照表"]
head = [str(c.value or "") for c in ws[1]]
check("① 表头有「计划ID」和填写列",
      "计划ID" in head and "项目编号（请填这里）" in head, f"{head}")
n_plans = ws.max_row - 1
check("① 导出了本科室的计划", n_plans > 0, f"{n_plans} 条")

# 只导本科室的
from services.dept import dept_names
names_a = set(dept_names("SBK"))
i_dept = head.index("归口科室")
i_dm = head.index("需求科室")
alien = [ws.cell(i, 1).value for i in range(2, ws.max_row + 1)
         if str(ws.cell(i, i_dept + 1).value or "") not in names_a
         and str(ws.cell(i, i_dm + 1).value or "") not in names_a]
check("① 只导出本科室的计划", not alien, f"混入 {len(alien)} 条")

ws2 = wb["项目清单（查编号用）"]
check("① 项目清单有内容", ws2.max_row > 1, f"{ws2.max_row - 1} 个项目")
nums = [str(ws2.cell(i, 1).value or "") for i in range(2, ws2.max_row + 1)]
check("① 项目清单带编号", any(nums), f"示例 {nums[:2]}")

# 未关联的排前面
i_cur = head.index("当前已关联项目")
first_linked = next((i for i in range(2, ws.max_row + 1)
                     if str(ws.cell(i, i_cur + 1).value or "").strip()), None)
last_unlinked = max([i for i in range(2, ws.max_row + 1)
                     if not str(ws.cell(i, i_cur + 1).value or "").strip()] or [0])
check("① 未关联的排在前面", first_linked is None or last_unlinked < first_linked,
      f"最后一条未关联在第 {last_unlinked} 行，第一条已关联在第 {first_linked} 行")

# ═══ 二、填好导回来 ═════════════════════════════════════════════
print("\n── 二、填好导回 ──")
i_id, i_fill = head.index("计划ID"), head.index("项目编号（请填这里）")
plan_ids = [int(ws.cell(i, i_id + 1).value) for i in range(2, min(ws.max_row + 1, 8))]
# 本科室可见的项目编号
r = sa.get("/api/project-monitor/projects", query_string={"page": 1, "page_size": 10})
proj_rows = (r.get_json() or {}).get("data") or []
free = [p for p in proj_rows if p.get("number")]
check("② 有可用的项目编号", len(free) >= 2, f"{len(free)} 个")


def build(pairs, plan_override=None):
    """按 (计划ID, 项目编号) 造一张回传表。"""
    w = Workbook(); s_ = w.active; s_.title = "计划对照表"
    s_.append(head)
    for pid_, num in pairs:
        row = [""] * len(head)
        row[i_id] = pid_ if plan_override is None else plan_override
        row[i_fill] = num
        s_.append(row)
    b = _io.BytesIO(); w.save(b); b.seek(0)
    return b


# 预览
buf = build([(plan_ids[0], free[0]["number"])])
r = sa.post("/api/procurement-plans/link-import",
            data={"file": (buf, "t.xlsx"), "dry_run": "1"},
            content_type="multipart/form-data")
d = r.get_json() or {}
check("② 预览能跑", r.status_code == 200, f"HTTP {r.status_code} {str(d)[:80]}")
check("② 预览说可关联 1 条", d.get("will_link") == 1, f"{d.get('will_link')}")
check("② 预览不写库",
      db.session.get(ProcurementPlan, plan_ids[0]).project_id != free[0]["id"])

# 真写
buf = build([(plan_ids[0], free[0]["number"])])
r = sa.post("/api/procurement-plans/link-import",
            data={"file": (buf, "t.xlsx")}, content_type="multipart/form-data")
d = r.get_json() or {}
check("② 真导入成功", d.get("will_link") == 1, f"{d.get('message')}")
db.session.expire_all()
check("② 关联落库了",
      db.session.get(ProcurementPlan, plan_ids[0]).project_id == free[0]["id"])
check("② 记了是谁导的",
      bool(db.session.get(ProcurementPlan, plan_ids[0]).linked_by))

# 空着的不动
buf = build([(plan_ids[1], "")])
r = sa.post("/api/procurement-plans/link-import",
            data={"file": (buf, "t.xlsx")}, content_type="multipart/form-data")
d = r.get_json() or {}
check("② 空着的一律不动", d.get("will_link") == 0 and d.get("skipped") == 1,
      f"link={d.get('will_link')} skip={d.get('skipped')}")

# ═══ 三、填错了要拦住 ═══════════════════════════════════════════
print("\n── 三、填错要拦住 ──")
buf = build([(plan_ids[1], "根本不存在的编号XYZ")])
d = (sa.post("/api/procurement-plans/link-import", data={"file": (buf, "t.xlsx")},
             content_type="multipart/form-data").get_json() or {})
check("③ 编号不存在会报出来",
      d.get("will_link") == 0 and any("找不到项目编号" in e for e in d.get("errors") or []),
      f"{(d.get('errors') or [''])[0][:50]}")

# 同一个编号填了两行
buf = build([(plan_ids[1], free[1]["number"]), (plan_ids[2], free[1]["number"])])
d = (sa.post("/api/procurement-plans/link-import", data={"file": (buf, "t.xlsx")},
             content_type="multipart/form-data").get_json() or {})
check("③ 同一编号重复填只认第一条",
      d.get("will_link") == 1 and any("不止一次" in e for e in d.get("errors") or []),
      f"link={d.get('will_link')}")

# 项目已被别的计划占用
buf = build([(plan_ids[3], free[0]["number"])])
d = (sa.post("/api/procurement-plans/link-import", data={"file": (buf, "t.xlsx")},
             content_type="multipart/form-data").get_json() or {})
check("③ 项目已被占用会拦住",
      d.get("will_link") == 0 and any("已被计划" in e for e in d.get("errors") or []),
      f"{(d.get('errors') or [''])[0][:60]}")

# 别的科室的计划
buf = build([(plan_ids[0], free[0]["number"])])
d = (sb.post("/api/procurement-plans/link-import", data={"file": (buf, "t.xlsx")},
             content_type="multipart/form-data").get_json() or {})
check("③ 别科室导不动本科室的计划",
      d.get("will_link") == 0 and any("不属于本科室" in e for e in d.get("errors") or []),
      f"{(d.get('errors') or [''])[0][:50]}")

# 表头被改坏
w = Workbook(); s_ = w.active; s_.title = "计划对照表"; s_.append(["随便", "乱写"])
b = _io.BytesIO(); w.save(b); b.seek(0)
r = sa.post("/api/procurement-plans/link-import", data={"file": (b, "t.xlsx")},
            content_type="multipart/form-data")
check("③ 表头对不上会明确报错",
      r.status_code == 400 and "表头对不上" in (r.get_json() or {}).get("error", ""),
      f"HTTP {r.status_code}")

# 传个不是 xlsx 的
r = sa.post("/api/procurement-plans/link-import",
            data={"file": (_io.BytesIO(b"hello"), "t.txt")},
            content_type="multipart/form-data")
check("③ 非 xlsx 被拒", r.status_code == 400, f"HTTP {r.status_code}")

# ── 回滚 ──────────────────────────────────────────────────────
db.session.expire_all()
n = 0
for pid_, old in before.items():
    row = db.session.get(ProcurementPlan, pid_)
    if row is not None and row.project_id != old:
        row.project_id = old
        row.linked_by = ""
        row.linked_at = ""
        n += 1
db.session.commit()
print(f"\n已回滚 {n} 条关联")

print(f"通过 {ok} 项" + (f"，失败 {len(bad)}：{bad}" if bad else "，全部通过"))
sys.exit(1 if bad else 0)

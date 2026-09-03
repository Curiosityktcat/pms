"""考核表成稿：导出 Excel + 打印页，两个出口共用一份行模型。

版式照客户《2025年10月 版本2招标代理机构服务质量考核评价表》一比一抄，
四段（基本信息 / 考核内容及评分标准 / 一票否决 / 综合评价）加表末注 1-5。

硬要求：打印出来不能超过 2 页 A4——两页正好双面一张纸。
所以 Excel 里设了 fitToWidth=1 / fitToHeight=2 的缩放打印，
打印页用 @page A4 + 固定字号行高，实测 15 个评分项排下来一页半。

一份行模型 (build_rows) 出两种成稿，是为了以后改评分表只改一处，
不会出现「Excel 改了、打印页没改」这种两份不一样的事。
"""
import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services import agency_assessment as svc

CHECKED, UNCHECKED = "☑", "□"


def _fmt(n):
    """分数写成 +1.5 / -0.3 / 0 这样，一眼看得出加还是扣。"""
    if n is None or n == "":
        return ""
    v = float(n)
    if v == 0:
        return "0"
    txt = f"{v:g}"
    return f"+{txt}" if v > 0 else txt


def build_rows(row, items, meta_veto=None):
    """把一份考核表拍平成打印用的段落结构。

    row  —— AgencyAssessment.to_dict() 或等价 dict
    items —— svc.build_items() 的结果（带最终得分与备注）
    """
    veto_items = meta_veto or svc.VETO_ITEMS
    hit = set(row.get("veto") or [])
    total = svc.total_of(items)

    subj = [
        ("经办人响应的及时性", row.get("subj_timeliness") or ""),
        ("经办人水平和能力", row.get("subj_ability") or ""),
        ("经办人合作态度及协调能力", row.get("subj_attitude") or ""),
    ]
    return {
        "title": "招标代理机构服务质量考核评价表",
        "project": f"{row.get('project_name', '')}\n{row.get('project_number', '')}".strip(),
        "agency": row.get("agency_name", ""),
        "items": [
            {
                "name": i["name"],
                "standard": i["standard"],
                "score": _fmt(i.get("score")),
                "note": i.get("note", "") or "",
            } for i in items
        ],
        "total": f"{total:g}",
        "veto_lines": [
            f"{CHECKED if v['key'] in hit else UNCHECKED} （{'一二三四五六七八九'[n]}）{v['name']}"
            for n, v in enumerate(veto_items)
        ],
        "veto_hit": bool(hit),
        "veto_note": row.get("veto_note", "") or "",
        "subj": [
            (label, "".join(
                f"{CHECKED if val == o else UNCHECKED}{o}    " for o in svc.SUBJ_OPTIONS))
            for label, val in subj
        ],
        "comment": row.get("comment", "") or "",
        "assessor": row.get("assessor", "") or "",
        "assessed_at": (row.get("assessed_at", "") or "").replace("T", " ")[:10],
        "status": row.get("status", "草稿"),
        "footnotes": [
            "1.本考核表得分低于90分，将暂停下一轮代理机构项目的拟派，多个项目考核加分累计满10分的，提前一轮代理机构项目拟派。",
            "2.日常考核扣分有效期为3个月，在日常工作中累加计算，但按年计入年度综合考核。",
            "3.日常考核有效期内累计加扣分达30分的，暂停采购代理资格3个月，暂停期间禁止代理我院任何采购项目。整改或暂停期满，经采购部同意后，予以恢复。",
            "4.代理机构日常管理考核实行动态化考核管理，其结果作为限制一定时期内承接项目代理业务和一票否决等方面的主要依据。年度考核在每年年末或下年年初进行，作为下一年度内代理资格管理及后续选择代理机构的主要依据。",
            "5.代理机构在日常考核和年终综合考核中，发生所列“一票否决”行为，实行一票否决制，暂停代理我院采购项目资格一年，暂停期间禁止代理我院任何采购项目。",
        ],
    }


# ── Excel ─────────────────────────────────────────────────────────
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="F2F2F2")
FONT = "宋体"


def _cell(ws, r, c, value, *, bold=False, size=9, wrap=True,
          h="left", v="center", fill=None):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = Font(name=FONT, size=size, bold=bold)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    cell.border = BORDER
    if fill:
        cell.fill = fill
    return cell


def _band(ws, r, text, ncols=4, size=10):
    """整行占满的段标题（一、二、三、四 这几段）。"""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    _cell(ws, r, 1, text, bold=True, size=size, h="left", fill=HEAD_FILL)
    for c in range(2, ncols + 1):
        ws.cell(row=r, column=c).border = BORDER


def _merged_h(text, per_line=92, line_h=11.5, floor=20):
    """整行合并单元格的行高：Excel 对合并区不做自适应，只能按字数估。

    宁可估高留白，也不能估矮——估矮了文字被裁，打出来的考核表就少了内容。
    """
    n = 1
    for seg in str(text or "").split("\n"):
        n += max(1, (len(seg) + per_line - 1) // per_line) - 1
        n += 1
    return max(floor, (n - 1) * line_h + 6)


def to_xlsx(data):
    """按 build_rows 的结构出一份 .xlsx，返回字节流。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "考核评价表"

    widths = [46, 34, 9, 18]          # 考核内容 / 评分标准 / 扣分加分 / 备注
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _cell(ws, r, 1, data["title"], bold=True, size=14, h="center")
    ws.row_dimensions[r].height = 28
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = BORDER
    r += 1

    _band(ws, r, "一、基本信息"); r += 1
    _cell(ws, r, 1, "项目名称及编号：", bold=True, h="left")
    _cell(ws, r, 2, data["project"])
    _cell(ws, r, 3, "代理机构名称：", bold=True, h="center")
    _cell(ws, r, 4, data["agency"])
    ws.row_dimensions[r].height = max(32, 12.5 * (len(data["project"].split("\n")) + 1))
    r += 1

    _band(ws, r, "二、考核内容及评分标准"); r += 1

    for c, t in enumerate(["考核内容", "评分标准", "扣分/加分", "备注/说明"], start=1):
        _cell(ws, r, c, t, bold=True, h="center", fill=HEAD_FILL)
    ws.row_dimensions[r].height = 16
    r += 1

    for it in data["items"]:
        _cell(ws, r, 1, it["name"], size=8)
        _cell(ws, r, 2, it["standard"], size=8)
        _cell(ws, r, 3, it["score"], h="center", bold=True)
        _cell(ws, r, 4, it["note"], size=8)
        # 行高故意不写死：写死了备注一长就被裁掉，看不见的扣分理由等于没写。
        # 留空让 Excel/WPS 打开时按内容自适应，整表再靠 fitToHeight=2 缩进两页。
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    _cell(ws, r, 1, "本考核评价表满分100分，得分=满分100+以上各项扣分/加分之和。本考核表合计：",
          bold=True, size=9)
    ws.cell(row=r, column=2).border = BORDER
    _cell(ws, r, 3, data["total"], bold=True, size=11, h="center")
    _cell(ws, r, 4, "", h="center")
    ws.row_dimensions[r].height = 20
    r += 1

    _band(ws, r, "三、一票否决项目（代理机构有无以下行为）"); r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _cell(ws, r, 1, "\n".join(data["veto_lines"]), size=8, v="top")
    ws.row_dimensions[r].height = 10.5 * len(data["veto_lines"]) + 4
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = BORDER
    r += 1
    if data["veto_hit"]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        _cell(ws, r, 1, f"一票否决事实依据：{data['veto_note']}", size=8, v="top")
        for c in range(2, 5):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = _merged_h(
            f"一票否决事实依据：{data['veto_note']}", per_line=104, line_h=10.5, floor=18)
        r += 1

    _band(ws, r, "四、综合评价（非评分项）"); r += 1
    for label, marks in data["subj"]:
        _cell(ws, r, 1, label, size=9)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        _cell(ws, r, 2, marks, size=9)
        for c in range(3, 5):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 16
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _cell(ws, r, 1, f"建议或意见：{data['comment']}", size=9, v="top")
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = _merged_h(f"建议或意见：{data['comment']}", floor=34)
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _cell(ws, r, 1,
          f"采购部对接人签字：{data['assessor']}　　　　　　　　　　日期：{data['assessed_at'] or '　　年　　月　　日'}",
          size=9)
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 26
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _cell(ws, r, 1, "注：" + "\n".join(data["footnotes"]), size=7, v="top")
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 96
    last = r

    # ── 打印设置：A4 纵向、缩放到「宽 1 页 × 高 2 页」，最多两页正好双面 ──
    ws.print_area = f"A1:D{last}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 2
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    ws.page_margins.header = ws.page_margins.footer = 0.2
    ws.print_options.horizontalCentered = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── 打印页 ─────────────────────────────────────────────────────────
def _esc(t):
    return (str(t or "").replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace("\n", "<br>"))


def to_print_html(data, auto_print=True):
    """一张自带样式的打印页，浏览器直接 Ctrl+P，版心控制在 2 页 A4 内。"""
    items = "".join(
        f"<tr><td class=nm>{_esc(i['name'])}</td>"
        f"<td class=std>{_esc(i['standard'])}</td>"
        f"<td class=sc>{_esc(i['score'])}</td>"
        f"<td class=nt>{_esc(i['note'])}</td></tr>"
        for i in data["items"])
    vetos = "".join(f"<div class=v>{_esc(v)}</div>" for v in data["veto_lines"])
    veto_note = (f"<div class=vn>一票否决事实依据：{_esc(data['veto_note'])}</div>"
                 if data["veto_hit"] else "")
    subj = "".join(
        f"<tr><td class=sl>{_esc(label)}</td><td class=sm>{_esc(marks)}</td></tr>"
        for label, marks in data["subj"])
    notes = "".join(f"<div>{_esc(n)}</div>" for n in data["footnotes"])
    # 「最多两页」是硬指标，不能只靠版式估算：备注写长了照样撑到第三页。
    # 所以印之前先量一次实际高度，超了就整体缩到刚好两页（最多缩到 72%，
    # 再小就看不清了，那种情况说明备注写得实在太长，宁可让它溢出提醒人删）。
    fit = """<script>
(function(){
  function fit(){
    var mm = document.createElement('div');
    mm.style.cssText='height:100mm;position:absolute;visibility:hidden';
    document.body.appendChild(mm);
    var pxPerMm = mm.offsetHeight / 100; mm.remove();
    var limit = pxPerMm * (297 - 20) * 2;          // 两页 A4 去掉上下各 10mm 页边距
    var w = document.querySelector('.wrap');
    for (var z = 1; z > 0.72 && w.scrollHeight * z > limit; z -= 0.02) {}
    if (z < 1) { w.style.zoom = z.toFixed(2); }
  }
  window.addEventListener('load', function(){ fit(); AUTOPRINT });
})();
</script>"""
    auto = fit.replace("AUTOPRINT", "setTimeout(function(){window.print()}, 120);"
                       if auto_print else "")
    draft = ("<div class=draft>草稿 · 未提交</div>" if data["status"] != "已提交" else "")

    return f"""<!doctype html><html lang=zh-CN><head><meta charset=utf-8>
<title>{_esc(data['title'])}</title>
<style>
  /* 两页 A4 是硬指标：页边距压到 10mm，正文 8.5pt，15 个评分项刚好一页半 */
  @page {{ size: A4 portrait; margin: 10mm 9mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: "宋体", SimSun, serif; font-size: 8.5pt; line-height: 1.35;
         margin: 0; color: #000; }}
  .wrap {{ width: 192mm; margin: 0 auto; }}
  h1 {{ font-size: 15pt; text-align: center; margin: 0 0 6px; font-weight: 700;
       font-family: "黑体", SimHei, sans-serif; }}
  table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
  td, th {{ border: 0.5pt solid #000; padding: 2px 4px; vertical-align: middle;
           word-break: break-all; }}
  .band td {{ background: #f2f2f2; font-weight: 700; font-size: 9.5pt;
             font-family: "黑体", SimHei, sans-serif; }}
  thead th {{ background: #f2f2f2; font-weight: 700; text-align: center; }}
  .nm {{ width: 44%; }} .std {{ width: 32%; font-size: 7.8pt; color: #222; }}
  .sc {{ width: 8%; text-align: center; font-weight: 700; }}
  .nt {{ width: 16%; font-size: 7.8pt; }}
  .total td {{ font-weight: 700; }}
  .v {{ font-size: 8pt; line-height: 1.5; }}
  .vn {{ font-size: 8pt; margin-top: 3px; }}
  .sl {{ width: 30%; }} .sm {{ letter-spacing: 1px; }}
  .cm {{ min-height: 40px; vertical-align: top; }}
  .sign {{ padding: 8px 4px; }}
  .notes {{ font-size: 7pt; line-height: 1.4; color: #222; }}
  .draft {{ position: fixed; top: 4mm; right: 4mm; font-size: 9pt; color: #c00;
           border: 1pt solid #c00; padding: 1px 6px; }}
  /* 段落不要被切在两页中间，切了看着像少了内容 */
  tr, .v {{ page-break-inside: avoid; }}
  @media screen {{ body {{ background: #f5f5f5; padding: 16px; }}
    .wrap {{ background: #fff; padding: 12mm 9mm; box-shadow: 0 1px 6px rgba(0,0,0,.2); }} }}
</style></head><body>{draft}
<div class=wrap>
<h1>{_esc(data['title'])}</h1>
<table>
  <tr class=band><td colspan=4>一、基本信息</td></tr>
  <tr><td style="width:18%">项目名称及编号：</td><td style="width:42%">{_esc(data['project'])}</td>
      <td style="width:16%">代理机构名称：</td><td>{_esc(data['agency'])}</td></tr>
</table>
<table style="margin-top:-1px">
  <tr class=band><td colspan=4>二、考核内容及评分标准</td></tr>
  <thead><tr><th class=nm>考核内容</th><th class=std>评分标准</th>
    <th class=sc>扣分/加分</th><th class=nt>备注/说明</th></tr></thead>
  <tbody>{items}
  <tr class=total><td colspan=2>本考核评价表满分100分，得分=满分100+以上各项扣分/加分之和。本考核表合计：</td>
    <td class=sc style="font-size:11pt">{_esc(data['total'])}</td><td></td></tr></tbody>
</table>
<table style="margin-top:-1px">
  <tr class=band><td>三、一票否决项目（代理机构有无以下行为）</td></tr>
  <tr><td>{vetos}{veto_note}</td></tr>
</table>
<table style="margin-top:-1px">
  <tr class=band><td colspan=2>四、综合评价（非评分项）</td></tr>
  {subj}
  <tr><td colspan=2 class=cm>建议或意见：{_esc(data['comment'])}</td></tr>
  <tr><td colspan=2 class=sign>采购部对接人签字：{_esc(data['assessor'])}
      &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
      日期：{_esc(data['assessed_at'] or '　　年　　月　　日')}</td></tr>
  <tr><td colspan=2 class=notes><b>注：</b>{notes}</td></tr>
</table>
</div>{auto}</body></html>"""


def filename_for(row):
    """下载文件名：机构 + 项目 + 日期，落到桌面上一眼认得出是哪份。"""
    day = (row.get("assessed_at") or "")[:10] or datetime.date.today().isoformat()
    name = (row.get("project_name") or "考核表")[:40]
    return f"服务质量考核表_{row.get('agency_name', '')}_{name}_{day}.xlsx"

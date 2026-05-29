"""
采购需求 Excel 模板生成 & 解析
支持五种需求类型：gov / competition / sole_source / inquiry / emergency
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 样式常量 ────────────────────────────────────────────────────
_BLUE_FILL  = PatternFill("solid", fgColor="2B579A")
_GRAY_FILL  = PatternFill("solid", fgColor="F2F2F2")
_LBLUE_FILL = PatternFill("solid", fgColor="EBF3FB")
_GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

_WHITE_BOLD = Font(color="FFFFFF", bold=True, size=10)
_BOLD10     = Font(bold=True, size=10)
_GRAY_TIP   = Font(color="888888", italic=True, size=9)
_NORMAL     = Font(size=10)

_WRAP  = Alignment(wrap_text=True, vertical="top")
_CENT  = Alignment(horizontal="center", vertical="center")

_THIN_BORDER = Border(
    left=Side(style="thin",   color="BFBFBF"),
    right=Side(style="thin",  color="BFBFBF"),
    top=Side(style="thin",    color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _cell(ws, row, col, value="", fill=None, font=None, align=None, border=True, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill      = fill
    if font:   c.font      = font
    if align:  c.alignment = align
    if border: c.border    = _THIN_BORDER
    if number_format: c.number_format = number_format
    return c


# ── 每种类型的基本信息字段定义 ────────────────────────────────────
# (label, field_name, example_value, note)
_BASIC_FIELDS: dict[str, list[tuple]] = {
    "gov": [
        ("采购需求名称",   "project_name",    "XX设备政府采购需求",             "必填"),
        ("需求科室",       "demand_dept",     "骨科",                           "必填"),
        ("归口管理科室",   "manage_dept",     "医学装备部",                     ""),
        ("采购年度",       "year",            "2026年",                         ""),
        ("预算金额（元）", "budget_amount",   100000,                           "填数字"),
        ("项目分类",       "category",        "货物",                           "货物/服务/工程 三选一"),
        ("采购组织形式",   "org_form",        "分散采购",                       "集中采购/分散采购"),
        ("预算采购方式",   "budget_method",   "竞争性谈判",                     "公开招标/邀请招标/竞争性谈判/单一来源/询价/竞争性磋商"),
        ("项目概况",       "project_overview","XX设备用于…",                   "可填多行"),
        ("需求调查情况",   "survey_content",  "已调研市场价格，参考供应商…",    "可填多行"),
    ],
    "competition": [
        ("采购需求名称",   "project_name",    "XX耗材院内竞选采购需求",          "必填"),
        ("需求科室",       "demand_dept",     "骨科",                            "必填"),
        ("归口管理科室",   "manage_dept",     "医学装备部",                      ""),
        ("采购年度",       "year",            "2026年",                          ""),
        ("预算金额（元）", "budget_amount",   80000,                             "填数字"),
        ("项目分类",       "category",        "货物",                            "货物/服务/工程 三选一"),
        ("项目概况",       "project_overview","该耗材用于…",                    "可填多行"),
    ],
    "sole_source": [
        ("采购需求名称",     "project_name",       "XX单一来源采购需求",              "必填"),
        ("需求科室",         "demand_dept",         "骨科",                           "必填"),
        ("归口管理科室",     "manage_dept",         "医学装备部",                     ""),
        ("采购年度",         "year",                "2026年",                         ""),
        ("预算金额（元）",   "budget_amount",       60000,                            "填数字"),
        ("项目分类",         "category",            "货物",                           "货物/服务/工程 三选一"),
        ("项目概况",         "project_overview",    "该产品用于…",                   "可填多行"),
        ("单一来源论证理由", "sole_source_reason",  "该产品具有唯一性，原因为…",     "必填，说明不可替代性"),
    ],
    "inquiry": [
        ("采购需求名称",             "project_name",               "2025年XX服务采购项目",    "必填"),
        ("需求科室",                 "demand_dept",                 "审计科",                  "必填"),
        ("归口管理科室",             "manage_dept",                 "医学装备部",              ""),
        ("采购年度",                 "year",                        "2026年",                  ""),
        ("预算金额（元）",           "budget_amount",               30000,                     "填数字"),
        ("项目分类",                 "category",                    "服务",                    "货物/服务/工程 三选一"),
        ("采购方式",                 "procurement_method",          "院内询价",                "院内询价(2-5万)/院内议价(2万以下)"),
        ("特定资格要求",             "qualification_requirements",  "供应商需具备…证书",       "通用条款已固定，此处填写特定要求"),
        ("技术参数/服务要求",        "tech_requirements",
            "1. 项目内容\n1.1 …\n\n2. 具体要求\n2.1 …",          "必填，详细描述项目内容及要求"),
        ("付款方式",                 "payment_terms",
            "验收合格并提供发票后20个工作日内支付100%",            ""),
        ("售后服务要求",             "inq_after_sales",
            "自合同起始日起提供1年相关咨询服务",                   ""),
        ("交付时间",                 "inq_delivery_time",
            "合同签订后2个月内交付",                               ""),
        ("服务/交付地点",            "inq_delivery_location",
            "内江市汉安大道西段1866号",                            ""),
        ("验收标准",                 "acceptance_standard",
            "完成成果，经采购人审核无异议后完成验收",              ""),
        ("履约保证金",               "inq_performance_bond",       "不收取",                  ""),
        ("违约责任",                 "breach_terms",               "无",                      ""),
        ("其他要求",                 "inq_other_requirements",
            "（1）报价含所有费用…\n（2）…",                       "可多行填写"),
    ],
}

_EMERGENCY_FIELDS = [
    ("申请名称（列表显示用）",       "project_name",         "骨科急用接骨板紧急采购申请",  "必填"),
    ("申请科室",                     "demand_dept",           "骨科",                        "必填"),
    ("预计使用紧急耗材时间",         "expected_use_time",     "2026-06-10",                  "格式 YYYY-MM-DD"),
    ("耗材通用名",                   "consumable_name",       "骨科接骨板",                  "必填"),
    ("预算单价（元）",               "consumable_unit_price", 5000,                          "填数字"),
    ("单位",                         "consumable_unit",       "套",                          ""),
    ("申请数量",                     "apply_quantity",        2,                             "填整数"),
    ("生产厂家",                     "manufacturer",          "XX医疗器械有限公司",          ""),
    ("型号",                         "model_spec",            "A型",                         ""),
    ("用途",                         "purpose_type",          "手术",                        "手术/治疗/检查 三选一"),
    ("手术名称（用途=手术时填）",    "surgery_name",          "股骨干骨折内固定术",          ""),
    ("手术级别",                     "surgery_level",         "三级手术",                    "一/二/三/四级手术"),
    ("耗材使用证（注册证号）",       "license_number",        "国械注准20xxXXXXXX",          ""),
    ("院内有无类似产品",             "has_similar_product",   "无",                          "有/无 二选一"),
    ("是否需要配套设备/器械",        "needs_companion",       "无需",                        "需要/无需 二选一"),
    ("进口/国产",                    "import_domestic",       "国产",                        "进口/国产 二选一"),
    ("产品类型",                     "product_on_catalog",    "非挂网产品",                  "挂网产品/非挂网产品 二选一"),
    ("适用范围（填数字1-4）",        "applicable_scope",      "3",
        "1=直接严重影响教学科研临床；2=病人就医突发事件；3=危急重症手术及抢救；4=其他"),
    ("申购耗材用途",                 "clinical_use_desc",
        "该耗材用于骨折固定手术，具有…优势，适应症为…",          "必填，详述临床用途、优势、适应症"),
]

# 标的清单列定义 (label, field_key, width, example1, example2)
_ITEMS_COLS = [
    ("品目编号",       "item_no",      10, "1-1",    "1-2"),
    ("品目类别",       "category",     14, "医疗设备","医疗耗材"),
    ("标的名称",       "name",         22, "骨科接骨板","缝合线"),
    ("预算单价（元）", "unit_price",   14, 5000,      120),
    ("数量",           "quantity",     8,  2,          50),
    ("单位",           "unit",         8,  "套",       "包"),
    ("技术要求摘要",   "requirements", 28, "材质：钛合金；符合YY标准", ""),
]
# 金额列（只读，由前端计算）
_AMOUNT_COL_LABEL = "预算金额（自动计算）"


def _set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


def _add_instruction_sheet(wb, demand_type: str):
    """添加「填写说明」Sheet"""
    ws = wb.create_sheet("填写说明", 0)
    ws.column_dimensions["A"].width = 80

    type_names = {
        "gov": "政府采购", "competition": "院内竞选", "sole_source": "单一来源",
        "inquiry": "询议价", "emergency": "紧急采购"
    }
    title = f"内江市第一人民医院——{type_names.get(demand_type,'采购')}需求 Excel 导入模板使用说明"

    _cell(ws, 1, 1, title, fill=_BLUE_FILL, font=_WHITE_BOLD, align=_CENT, border=False)
    ws.row_dimensions[1].height = 28

    instructions = [
        "【使用步骤】",
        "① 在对应 Sheet 中填写数据（蓝色标注区域）",
        "② 基本信息 Sheet：在「请填写（C列）」中输入值，不要修改 A、B 列",
        "③ 标的清单 Sheet：从第 3 行开始逐行填写，示例行（第2行灰色）可删除或保留",
        "④ 填写完毕后保存 Excel（.xlsx 格式），回到系统点击「从 Excel 导入」上传此文件",
        "⑤ 系统自动填充表单，导入后请核查后再保存",
        "",
        "【注意事项】",
        "• 请勿修改 Sheet 名称，否则导入失败",
        "• 请勿修改「基本信息」Sheet 中 A 列和 B 列的内容",
        "• 预算金额、数量、单价等数值列请填写纯数字（不要加单位或千位符）",
        "• 若某字段不需要填写，可以留空",
        "• 「标的清单」中「预算金额」列由系统自动计算，无需手动填写",
    ]
    for i, line in enumerate(instructions, start=2):
        c = ws.cell(row=i, column=1, value=line)
        c.font = _BOLD10 if line.startswith("【") else (_GRAY_TIP if line == "" else _NORMAL)
        c.alignment = _WRAP
        ws.row_dimensions[i].height = 18
    ws.sheet_view.showGridLines = False


def _add_basic_sheet(wb, fields: list[tuple], sheet_name="基本信息"):
    """添加基本信息 Sheet（key-value 格式）"""
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 48

    # 表头行
    headers = ["字段名称", "填写示例 / 说明", "请填写（导入时读取此列）"]
    for ci, h in enumerate(headers, 1):
        _cell(ws, 1, ci, h, fill=_BLUE_FILL, font=_WHITE_BOLD, align=_CENT)
    ws.row_dimensions[1].height = 22

    for ri, (label, _field, example, note) in enumerate(fields, start=2):
        hint = f"{example}  ← 示例"
        if note:
            hint += f"（{note}）"
        _cell(ws, ri, 1, label,  fill=_GRAY_FILL, font=_BOLD10, align=_WRAP)
        _cell(ws, ri, 2, hint,   fill=_LBLUE_FILL, font=_GRAY_TIP, align=_WRAP)
        _cell(ws, ri, 3, "",     fill=_WHITE_FILL, font=_NORMAL,   align=_WRAP)
        ws.row_dimensions[ri].height = max(
            18, min(80, str(example).count('\n') * 16 + 20)
        )
    ws.sheet_view.showGridLines = False
    return ws


def _add_items_sheet(wb, sheet_name="标的清单"):
    """添加标的清单 Sheet"""
    ws = wb.create_sheet(sheet_name)

    # 列宽
    for ci, (_, _, w, _, _) in enumerate(_ITEMS_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    # 金额列
    amt_col = len(_ITEMS_COLS) + 1
    ws.column_dimensions[get_column_letter(amt_col)].width = 16

    # 标题行
    for ci, (label, _, _, _, _) in enumerate(_ITEMS_COLS, 1):
        _cell(ws, 1, ci, label, fill=_BLUE_FILL, font=_WHITE_BOLD, align=_CENT)
    _cell(ws, 1, amt_col, _AMOUNT_COL_LABEL, fill=_BLUE_FILL, font=_WHITE_BOLD, align=_CENT)
    ws.row_dimensions[1].height = 24

    # 示例行（灰色）
    for ci, (_, _, _, ex1, _) in enumerate(_ITEMS_COLS, 1):
        _cell(ws, 2, ci, ex1, fill=_GRAY_FILL, font=_GRAY_TIP, align=_WRAP)
    _cell(ws, 2, amt_col, "=E2*F2", fill=_GRAY_FILL, font=_GRAY_TIP)
    ws.row_dimensions[2].height = 18

    # 预留 30 行可填写区（带浅色底）
    for ri in range(3, 33):
        for ci in range(1, len(_ITEMS_COLS) + 1):
            _cell(ws, ri, ci, "", fill=_WHITE_FILL, font=_NORMAL, align=_WRAP)
        # 金额公式
        c = ws.cell(row=ri, column=amt_col)
        c.value = f"=E{ri}*F{ri}"
        c.fill   = _LBLUE_FILL
        c.font   = _GRAY_TIP
        c.border = _THIN_BORDER
        ws.row_dimensions[ri].height = 18

    ws.sheet_view.showGridLines = False
    return ws


# ── 公开接口 ────────────────────────────────────────────────────

def generate_template(demand_type: str) -> tuple[io.BytesIO, str]:
    """生成 Excel 模板，返回 (BytesIO, filename)"""
    wb = Workbook()
    # 删除默认 Sheet
    del wb[wb.sheetnames[0]]

    _add_instruction_sheet(wb, demand_type)

    if demand_type == "emergency":
        _add_basic_sheet(wb, _EMERGENCY_FIELDS, sheet_name="申请信息")
    else:
        fields = _BASIC_FIELDS.get(demand_type, _BASIC_FIELDS["competition"])
        _add_basic_sheet(wb, fields, sheet_name="基本信息")
        _add_items_sheet(wb, sheet_name="标的清单")

    type_cn = {
        "gov": "政府采购", "competition": "院内竞选", "sole_source": "单一来源",
        "inquiry": "询议价", "emergency": "紧急采购"
    }.get(demand_type, demand_type)
    filename = f"内江一院_{type_cn}需求导入模板.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, filename


def parse_excel(demand_type: str, file_bytes: bytes) -> dict:
    """
    解析上传的 Excel，返回：
    {
      "basic_data": { field_name: value, ... },
      "items":      [ { item_no, category, name, ... }, ... ]
    }
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    basic_data: dict = {}
    items: list[dict] = []

    # ── 解析基本信息 ────────────────────────────────────────────
    if demand_type == "emergency":
        sheet_name, field_defs = "申请信息", _EMERGENCY_FIELDS
    else:
        sheet_name, field_defs = "基本信息", _BASIC_FIELDS.get(demand_type, _BASIC_FIELDS["competition"])

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 建立 label → field_name 映射
        label_map = {label: fname for label, fname, *_ in field_defs}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip()
            value = row[2] if len(row) > 2 else None  # C 列
            if value is None or str(value).strip() == "":
                continue
            fname = label_map.get(label)
            if fname:
                # 数值类型转换
                if fname in ("budget_amount", "consumable_unit_price", "eval_price_score"):
                    try:
                        basic_data[fname] = float(value)
                    except (ValueError, TypeError):
                        basic_data[fname] = value
                elif fname == "apply_quantity":
                    try:
                        basic_data[fname] = int(value)
                    except (ValueError, TypeError):
                        basic_data[fname] = value
                else:
                    basic_data[fname] = str(value).strip()

    # ── 解析标的清单 ────────────────────────────────────────────
    if demand_type != "emergency" and "标的清单" in wb.sheetnames:
        ws = wb["标的清单"]
        fields_order = [f for _, f, *_ in _ITEMS_COLS]  # 按列顺序
        # min_row=3：第1行=表头，第2行=灰色示例行（跳过）
        for row in ws.iter_rows(min_row=3, values_only=True):
            # 跳过示例行（第2行）和空行
            if not row or all(v is None for v in row):
                continue
            name_val = row[2] if len(row) > 2 else None  # 标的名称在第3列
            if name_val is None or str(name_val).strip() == "":
                continue
            item: dict = {}
            for ci, fname in enumerate(fields_order):
                raw = row[ci] if ci < len(row) else None
                if fname in ("unit_price",):
                    try:
                        item[fname] = float(raw or 0)
                    except (ValueError, TypeError):
                        item[fname] = 0.0
                elif fname == "quantity":
                    try:
                        item[fname] = int(float(raw or 0))
                    except (ValueError, TypeError):
                        item[fname] = 0
                else:
                    item[fname] = str(raw).strip() if raw is not None else ""
            # 计算金额
            item["amount"] = round(item.get("unit_price", 0) * item.get("quantity", 0), 2)
            items.append(item)

    return {"basic_data": basic_data, "items": items}

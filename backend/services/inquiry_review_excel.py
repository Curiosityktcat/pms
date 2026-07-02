"""评定标报告 Excel 生成（询议价评审，模块8.1）

以用户模板 医院模板/文件汇总/4.评定标报告模板（询议价）.xlsx 的第一个 sheet 为版式，
项目每轮评审生成一个 sheet（第一轮叫「评定标报告」，之后「评定标报告（第N次）」），
最后保留「附件三、供应商名单」并按最新一轮填写有效/无效供应商。

版式基准（供应商 3 行：第 9-11 行）：
  1 标题 / 2 一、项目信息 / 3 项目名称 / 4 采购内容 / 5 预算金额+编号+包号
  6 采购时间+地点+评审办法 / 7 二、评审过程 / 8 表头 / 9-11 供应商
  12 采购结果 / 13 备注 / 14 签字 / 15 四、确认采购结果 / 16 确认
  17 四、评审地点和评审委员会成员名单 / 18 评审地点 / 19 名单 / 20 五、开标情况 / 21 正文

注意：openpyxl 3.1 的 insert_rows/delete_rows 只移动单元格值，
**不**同步合并单元格与行高，必须手动修正（_shift_merges/_shift_heights）。
"""
import copy
import datetime
import os
from io import BytesIO

import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.cell_range import CellRange

from models import db
from models.inquiry_letter import InquiryLetter
from models.inquiry_supplier import InquirySupplier
from models.inquiry_review import InquiryReview

HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(
    HERE, "..", "..",
    "医院模板", "文件汇总", "4.评定标报告模板（询议价）.xlsx",
)
BASE_SHEET = "评定标报告（第一次有效供应商满足三家）"
APPENDIX_SHEET = "附件三、供应商名单"

_CN = ['', '一', '二', '三', '四', '五', '六', '七', '八', '九', '十']


def _cn_ord(n: int) -> str:
    if n < len(_CN):
        return _CN[n]
    if n < 20:
        return "十" + _CN[n - 10]
    return str(n)


def _row_op(ws, at: int, count: int, insert: bool):
    """带合并单元格/行高修正的插行（insert=True，在 at 前插 count 行）或
    删行（insert=False，删 [at, at+count) ）。

    openpyxl 3.1 的 insert_rows/delete_rows 只移动单元格值，不处理合并区与行高；
    且删行后旧合并区的占位 MergedCell 已不在，unmerge 会 KeyError。
    因此顺序必须是：先全部解除合并（此时单元格都在）→ 行操作 → 按新位置重新合并。"""
    delta = count if insert else -count
    old_ranges = [str(rng) for rng in ws.merged_cells.ranges]
    new_ranges = []
    for rng in old_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        if not insert and min_row >= at and max_row < at + count:
            continue  # 整体在被删行内，丢弃
        threshold = at if insert else at + count
        if min_row >= threshold:
            min_row += delta
            max_row += delta
        new_ranges.append(str(CellRange(min_col=min_col, min_row=min_row,
                                        max_col=max_col, max_row=max_row)))

    heights = {r: dim.height for r, dim in ws.row_dimensions.items()
               if dim.height is not None}

    for rng in old_ranges:
        ws.unmerge_cells(rng)
    if insert:
        ws.insert_rows(at, count)
    else:
        ws.delete_rows(at, count)
    for rng in new_ranges:
        ws.merge_cells(rng)

    for r in list(heights):
        ws.row_dimensions[r].height = None
    threshold = at if insert else at + count
    for r, h in heights.items():
        if not insert and at <= r < at + count:
            continue  # 被删行的行高丢弃
        nr = r + delta if r >= threshold else r
        ws.row_dimensions[nr].height = h


def _set(ws, coord: str, value):
    ws[coord] = value


def _adjust_supplier_rows(ws, n: int, first: int = 9, base: int = 3,
                          name_merge=("B", "C"), row_height: float = 60.0):
    """把从 first 行起的 base 个数据行调整为 n 行，返回行数变化 delta。"""
    n = max(n, 1)
    delta = n - base
    if delta > 0:
        insert_at = first + base
        _row_op(ws, insert_at, delta, insert=True)
        for i in range(delta):
            r = insert_at + i
            for col in "ABCDEFGHIJK":
                src = ws[f"{col}{first}"]
                dst = ws[f"{col}{r}"]
                dst._style = copy.copy(src._style)
            if name_merge:
                ws.merge_cells(f"{name_merge[0]}{r}:{name_merge[1]}{r}")
            ws.row_dimensions[r].height = row_height
    elif delta < 0:
        _row_op(ws, first + n, -delta, insert=False)
    return delta


def _parse_date(s: str):
    try:
        return datetime.datetime.strptime((s or "")[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _submitted(s: InquirySupplier) -> bool:
    """是否递交了响应（只有已响应的供应商才出现在评定标报告里）"""
    return bool(s.responded)


def _fmt_amount(v) -> str:
    if v is None:
        return ""
    return f"{v:g}元" if isinstance(v, float) else f"{v}元"


def _fill_round_sheet(ws, review: InquiryReview, sups):
    # 只列已响应的供应商；无人响应时留一行空行（不再把受邀名单全列成"/"）
    rows = [s for s in sups if _submitted(s)]
    n = max(len(rows), 1)
    _adjust_supplier_rows(ws, n)

    rnd = review.round_no or 1
    name = review.project_name or ""
    if rnd > 1:
        name = f"{name}（第{_cn_ord(rnd)}次）"
    _set(ws, "C3", name)
    _set(ws, "C4", review.content or "")
    _set(ws, "C5", review.budget or "")
    _set(ws, "F5", review.project_number or "")
    _set(ws, "H5", review.package_no or "/")
    d = _parse_date(review.review_date)
    _set(ws, "C6", d if d else (review.review_date or ""))
    _set(ws, "F6", review.location or "审计科")
    _set(ws, "H6", review.method or "")

    for i in range(n):
        r = 9 + i
        if i < len(rows):
            s = rows[i]
            _set(ws, f"A{r}", i + 1)
            _set(ws, f"B{r}", s.supplier_name or "")
            _set(ws, f"D{r}", _fmt_amount(s.quote_amount) or "/")
            _set(ws, f"E{r}", s.qual_pass or "/")
            _set(ws, f"F{r}", s.conform_pass or "/")
            _set(ws, f"G{r}", s.final_price or "/")
            _set(ws, f"H{r}", s.review_rank or "/")
        else:
            for col in "ABDEFGH":
                _set(ws, f"{col}{r}", "")
            _set(ws, f"A{r}", i + 1)

    base_shift = n - 3
    _set(ws, f"A{12 + base_shift}", f"采购结果：{review.result_text or ''}")
    remark = (review.remark or "").strip()
    _set(ws, f"A{13 + base_shift}", f"备注：{remark}" if remark else "备注")
    _set(ws, f"A{18 + base_shift}", f"评审地点：{review.review_place or ''}")
    _set(ws, f"A{19 + base_shift}", f"评审委员会成员名单：{review.committee or ''}")
    if (review.bid_open_info or "").strip():
        _set(ws, f"A{21 + base_shift}", review.bid_open_info)


def _fill_appendix(ws, sups):
    """附件三：有效供应商名单（第3-5行，3格）/ 无效供应商名单（第8-13行，6格）。
    超出格数时插行扩展。"""
    submitted = [s for s in sups if _submitted(s)]
    valid = [s for s in submitted if s.qual_pass == "通过" and s.conform_pass == "通过"]
    invalid = [s for s in submitted if s not in valid]

    # 先扩无效区（在下方，先动不影响有效区行号）
    if len(invalid) > 6:
        extra = len(invalid) - 6
        _row_op(ws, 14, extra, insert=True)
        for i in range(extra):
            r = 14 + i
            for col in "ABC":
                ws[f"{col}{r}"]._style = copy.copy(ws[f"{col}13"]._style)
    if len(valid) > 3:
        extra = len(valid) - 3
        _row_op(ws, 6, extra, insert=True)
        for i in range(extra):
            r = 6 + i
            for col in "ABC":
                ws[f"{col}{r}"]._style = copy.copy(ws[f"{col}5"]._style)
            ws.merge_cells(f"B{r}:C{r}")

    for i, s in enumerate(valid):
        ws[f"A{3 + i}"] = i + 1
        ws[f"B{3 + i}"] = s.supplier_name or ""
    invalid_first = 8 + max(len(valid) - 3, 0)
    for i, s in enumerate(invalid):
        ws[f"A{invalid_first + i}"] = i + 1
        ws[f"B{invalid_first + i}"] = s.supplier_name or ""
        ws[f"C{invalid_first + i}"] = s.fail_reason or ""


def generate_review_excel(project_id: int, upto_inquiry_id: int = None) -> BytesIO:
    """生成项目的评定标报告工作簿：截至 upto_inquiry_id（含）的所有轮次各一个 sheet。"""
    letters = db.session.execute(
        db.select(InquiryLetter).filter_by(project_id=project_id)
        .order_by(InquiryLetter.id)
    ).scalars().all()
    if upto_inquiry_id:
        letters = [l for l in letters if l.id <= upto_inquiry_id]

    rounds = []  # [(round_no, review, suppliers)]
    for idx, letter in enumerate(letters, start=1):
        review = db.session.execute(
            db.select(InquiryReview).filter_by(inquiry_id=letter.id)
        ).scalar_one_or_none()
        if not review:
            continue
        sups = db.session.execute(
            db.select(InquirySupplier).filter_by(inquiry_id=letter.id)
            .order_by(InquirySupplier.id)
        ).scalars().all()
        rounds.append((review.round_no or idx, review, sups))

    if not rounds:
        raise ValueError("该项目尚无评审记录")

    wb = openpyxl.load_workbook(TEMPLATE)
    base = wb[BASE_SHEET]
    appendix = wb[APPENDIX_SHEET]
    template_sheets = [n for n in wb.sheetnames if n != APPENDIX_SHEET]

    for rnd, review, sups in rounds:
        ws = wb.copy_worksheet(base)
        ws.title = "评定标报告" if rnd <= 1 else f"评定标报告（第{_cn_ord(rnd)}次）"
        _fill_round_sheet(ws, review, sups)

    _fill_appendix(appendix, rounds[-1][2])

    # 删掉模板自带的全部场景示例 sheet，只留生成的轮次 + 附件三（移到最后）
    for name in template_sheets:
        del wb[name]
    wb.move_sheet(appendix, offset=len(wb.sheetnames))
    wb.active = 0

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# -*- coding: utf-8 -*-
"""把 8-14 那批科室账号补齐到 8-19 的标准，并给专班办公室建号。

背景：8-19 的批量建号只处理「还没账号的科室」，所以 8-14 先建的那 16 个账号
留在了旧状态——用曾用名当登录名、显示名没带负责人、而且 must_change_pw=0
（发下去的一次性密码至今有效且不强制改）。这个脚本把它们对齐。

改登录名是安全的：跑之前用 _business_refs 查过，这 16 个账号在业务数据里
零引用（建了没用过）。update_user 接口有意不允许改登录名，所以只能在这里做。

密码只写进 xlsx 和 600 权限的 txt，不打印、不进审计 detail。
"""
import os
import sys
import secrets
from datetime import datetime

sys.path.insert(0, os.path.expanduser("~/pms/backend"))

import app as A
from models import db
from models.dept import Dept
from models.user import User
from models.user_audit_log import UserAuditLog
from services.auth import hash_pw
from routes.user_admin_api import _dept_display, _dept_username, _dept_role, _password

# 8-14 那批（预保科/院办已停用且不在 8-15 名单里，不动）
OLD = ["人事科", "保卫科", "党办", "公共卫生科", "医务科", "医学装备部", "团委",
       "基建科", "审计科-科室", "宣传科", "总务科", "法务部", "科教科", "药剂科",
       "质控办", "检验科"]

ZB_CODE = "ZBB"
ZB_NAME = "新区医院后期建设项目专班办公室"
XCK_ALIAS = "宣传统战社工部（文明办）"

OUT_XLSX = "/home/huangxb/files/科室账号与初始密码_补发_2026-08-20.xlsx"
OUT_TXT = os.path.expanduser("~/pms/data/dept_accounts_0820.txt")


def main():
    a = A.create_app()
    rows = []
    with a.app_context():
        # ① 专班办公室进字典。不是编制科室，但医院授权它归口，dept_type=行后
        #    才能让 _dept_role 判成 dept_manage。
        zb = db.session.execute(db.select(Dept).filter_by(code=ZB_CODE)).scalar_one_or_none()
        if not zb:
            zb = Dept(code=ZB_CODE, name=ZB_NAME, aliases="专班办公室",
                      category="归口", dept_type="行后", head_name="",
                      sort_no=30, active=1,
                      note="医院授权的归口部门，非编制科室；负责人待补。2026-08-20 黄新博指示建号。")
            db.session.add(zb)
            db.session.flush()
            print("已建科室：%s (%s)" % (ZB_NAME, ZB_CODE))
        else:
            print("科室已存在：%s" % ZB_NAME)

        # ② 宣传科别名补全带后缀的写法，认领那 1 个项目
        xck = db.session.execute(db.select(Dept).filter_by(code="XCK")).scalar_one_or_none()
        if xck:
            al = [x for x in (xck.aliases or "").split(",") if x.strip()]
            if XCK_ALIAS not in al:
                al.append(XCK_ALIAS)
                xck.aliases = ",".join(al)
                print("已补别名：XCK += %s" % XCK_ALIAS)

        db.session.commit()

        # ③ 对齐 16 个老账号 + 给专班建号
        taken = set(db.session.execute(db.select(User.username)).scalars())
        targets = []
        for n in OLD:
            u = db.session.execute(db.select(User).filter_by(username=n)).scalar_one_or_none()
            if u:
                targets.append(u)
            else:
                print("!! 找不到账号 %s" % n)

        zb_user = db.session.execute(
            db.select(User).filter_by(dept_code=ZB_CODE)).scalar_one_or_none()

        for u in targets:
            dept = db.session.execute(
                db.select(Dept).filter_by(code=u.dept_code)).scalar_one_or_none()
            if not dept:
                print("!! %s 的科室码 %s 不在字典里，跳过" % (u.username, u.dept_code))
                continue
            old_name = u.username
            taken.discard(old_name)             # 自己占的名字不算冲突
            new_name = _dept_username(dept, taken)
            taken.add(new_name)
            u.username = new_name
            u.display_name = _dept_display(dept)
            u.role = _dept_role(dept)
            pw = _password()
            u.salt = secrets.token_hex(16)
            u.pw_hash = hash_pw(pw, u.salt)
            u.must_change_pw = 1
            db.session.add(UserAuditLog(
                actor="admin", actor_name="系统维护",
                action="update", target_id=u.id, target_username=new_name,
                detail=('{"source":"fix_dept_accounts_0820","before":"%s","after":"%s",'
                        '"reset_password":true,"must_change_pw":1}' % (old_name, new_name))))
            rows.append((dept.name, dept.head_name or "", new_name, pw,
                         "归口管理科室" if u.role == "dept_manage" else "需求科室",
                         dept.dept_type or "",
                         "改名+重置" if old_name != new_name else "重置"))

        if not zb_user:
            uname = _dept_username(zb, taken)
            taken.add(uname)
            pw = _password()
            salt = secrets.token_hex(16)
            zb_user = User(username=uname, display_name=_dept_display(zb),
                           role=_dept_role(zb), active=1, dept_code=zb.code,
                           agency_code="", salt=salt, pw_hash=hash_pw(pw, salt))
            zb_user.must_change_pw = 1
            db.session.add(zb_user)
            db.session.flush()
            db.session.add(UserAuditLog(
                actor="admin", actor_name="系统维护", action="create",
                target_id=zb_user.id, target_username=uname,
                detail='{"source":"fix_dept_accounts_0820","dept_code":"%s"}' % ZB_CODE))
            rows.append((zb.name, "", uname, pw, "归口管理科室", "行后", "新建"))

        db.session.commit()

    # ④ 出表
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "补发账号(%d)" % len(rows)
    ws.append(["科室", "负责人", "登录账号", "一次性密码", "角色", "科室类型", "本次处理"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(list(r))
    for i, w in enumerate([28, 10, 28, 16, 14, 10, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("怎么用")
    for line in [
        ["科室账号补发说明（2026-08-20）", ""],
        ["", ""],
        ["1", "这张表补的是 8-14 先建的那批账号，它们没进 8-19 那份表。"],
        ["", "  两张表合起来才是全部科室账号；8-19 那份里的密码不受影响、仍然有效。"],
        ["2", "本次这些账号做了三件事：登录名改成科室现用名、显示名补上负责人、"],
        ["", "  密码重置为一次性密码并打开强制改密（首登不改密码就什么都点不了）。"],
        ["3", "旧密码（8-14 发出去的那批）已全部失效，请按本表重新发放。"],
        ["4", "登录名变了的科室要特别通知，例如：党办→党委办公室、医务科→医务部、"],
        ["", "  质控办→医疗质量管理控制办公室、宣传科→宣传统战部。"],
        ["5", "新区医院后期建设项目专班办公室是本次新建，按医院授权给了归口管理科室权限；"],
        ["", "  它的负责人字段还空着，知道是谁之后在「用户管理」里补上显示名。"],
        ["", ""],
        ["注意", "这张表有密码，属于「我的文件库」私人区，别整份外传，按科室分别发。"],
    ]:
        ws2.append(line)
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 78
    wb.save(OUT_XLSX)

    with open(OUT_TXT, "w", encoding="utf-8") as f:
        f.write("PMS 科室账号补发 %s\n" % datetime.now().strftime("%Y-%m-%d %H:%M"))
        f.write("旧密码已失效。登录地址：内网 http://172.1.14.12:1573\n\n")
        for r in rows:
            f.write("%-28s %-16s %-14s %s\n" % (r[2], r[3], r[4], r[0]))
    os.chmod(OUT_TXT, 0o600)
    print("")
    print("处理 %d 个账号" % len(rows))
    print("已出表：%s" % OUT_XLSX)
    print("已出表：%s (600)" % OUT_TXT)


if __name__ == "__main__":
    main()
